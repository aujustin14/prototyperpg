import openpyxl
import pyperclip
import sys

currentLocation = "D:\MEGAsync Downloads\python rpg\python rpg.xlsx"
currentWorkbook = openpyxl.load_workbook(filename = currentLocation, data_only = True)

playerClasses = {
	0: "–",
	1: "Player A",
	2: "Player B",
	3: "Player C",
	# 1: "Paladin",
	# 2: "Bard",
	# 3: "Priest",
	# 4: "Ninja",
	# 5: "Archer",
	# 6: "Wizard"
}
playerClassesValues = list(playerClasses.values())

enemyNames = {
	0: "Empty",
	1: "Green Slime",
	2: "Blue Slime",
	3: "Zombie",
	4: "Demon Eye",
	5: "Blood Slime",
	6: "Flesh Slime",
	7: "Man Eater",
	8: "Demon Eye Bat",
	9: "Stone Slime",
	10: "Flesh Slime",
	11: "Man Eater",
	12: "Demon Bat",
}
enemyNamesValues = list(enemyNames.values())

bossNames = {
	0: "Empty",
	1: "Great Green Slime",
	2: "Great Blue Slime",
	3: "Armored Zombie",
	4: "Demon Eye Fleet",
	5: "Giant Blood Slime",
	6: "Giant Flesh Slime",
	7: "Stoic Man Eater",
	8: "Demon Bat Swarm",
}
bossNamesValues = list(bossNames.values())

skillTypes = {
	0: "–",
	1: "Attack",
	2: "Heal",
	3: "Revive",
	4: "Buff",
	5: "Debuff",
}

statTypes = {
	0: "–",
	1: "Melee",
	2: "Ranged"
}

elements = {
	0: "–",
	1: "Non-elemental",
	2: "Weapon-elemental",
	2: "Fire",
	3: "Water",
	4: "Ice",
}
elementsValues = list(elements.values())

targetingTypes = {
	0: "–",
	1: "Self",
	2: "Single Enemy",
	3: "All Enemies",
	4: "Single Alive Ally",
	5: "All Alive Allies",
	6: "Single Dead Ally",
	7: "All Dead Allies",
}

equipmentList = {
	0: "Empty-handed",
	1: "Naked",
	2: "Sword",
	3: "Chestplate",
	4: "Dagger",
	5: "Tunic",
	6: "Staff",
	7: "Robes",
}
equipmentValues = list(equipmentList.values())

skillsList = {
	0: ["Empty", "–"],
	1: ["Defend", "–"],
	2: ["Evade", "–"],
	3: ["Basic Attack A", "Player A"],
	4: ["Basic Attack B", "Player A"],
	5: ["Basic Attack C", "Player B"],
	6: ["Basic Attack D", "Player B"],
	7: ["Self-heal", "Player C"],
	8: ["Heal A", "Player C"],
	9: ["Heal B", "Player C"],
	10: ["Revive A", "Player C"],
	11: ["Revive B", "Player C"],
}
skillsListKeys = list(skillsList.keys())
skillsNameValues = []
skillsClassValues = []
for i in skillsListKeys:
	currentSkill = skillsList[i]
	skillsNameValues.append(currentSkill[0])
	skillsClassValues.append(currentSkill[1])

statusEffectsList = {
	0: "Healthy I",
	1: "Healthy II",
	2: "Healthy III",
	3: "Energetic I",
	4: "Energetic II",
	5: "Energetic III",
	6: "Strengthened I",
	7: "Strengthened II",
	8: "Strengthened III",
	9: "Sharpened I",
	10: "Sharpened II",
	11: "Sharpened III",
	12: "Shielded I",
	13: "Shielded II",
	14: "Shielded III",
	15: "Barriered I",
	16: "Barriered II",
	17: "Barriered III",
	18: "Accurate I",
	19: "Accurate II",
	20: "Accurate III",
	21: "Evasive I",
	22: "Evasive II",
	23: "Evasive III",
	24: "Resourceful I",
	25: "Resourceful II",
	26: "Resourceful III",
	27: "Offensive I",
	28: "Offensive II",
	29: "Offensive III",
	30: "Defensive I",
	31: "Defensive II",
	32: "Defensive III",
	33: "Agile I",
	34: "Agile II",
	35: "Agile III",
	36: "Enhanced I",
	37: "Enhanced II",
	38: "Enhanced III",
	39: "Enchanted I",
	40: "Enchanted II",
	41: "Enchanted III",
	42: "Empowered I",
	43: "Empowered II",
	44: "Empowered III",
	45: "Ill I",
	46: "Ill II",
	47: "Ill III",
	48: "Lethargic I",
	49: "Lethargic II",
	50: "Lethargic III",
	51: "Weakened I",
	52: "Weakened II",
	53: "Weakened III",
	54: "Dulled I",
	55: "Dulled II",
	56: "Dulled III",
	57: "Broken I",
	58: "Broken II",
	59: "Broken III",
	60: "Shattered I",
	61: "Shattered II",
	62: "Shattered III",
	63: "Blinded I",
	64: "Blinded II",
	65: "Blinded III",
	66: "Sluggish I",
	67: "Sluggish II",
	68: "Sluggish III",
	69: "Drained I",
	70: "Drained II",
	71: "Drained III",
	72: "Distracted I",
	73: "Distracted II",
	74: "Distracted III",
	75: "Offguard I",
	76: "Offguard II",
	77: "Offguard III",
	78: "Clumsy I",
	79: "Clumsy II",
	80: "Clumsy III",
	81: "Poisoned I",
	82: "Poisoned II",
	83: "Poisoned III",
	84: "Cursed I",
	85: "Cursed II",
	86: "Cursed III",
	87: "Empowered I",
	88: "Empowered II",
	89: "Empowered III",
	90: "Defending",
	91: "Evading",
	92: "Burn",
	93: "Blaze",
	94: "Scorch",
	95: "Wet",
	96: "Soak",
	97: "Drench",
	98: "Frozen",
	99: "Frostbite",
	100: "Frostburn",
	101: "Refreshing",
	102: "Regrowing",
	103: "Renewing",
	104: "Napping",
	105: "Resting",
	106: "Relaxing",
}
statusEffectsValues = list(statusEffectsList.values())

print("1. SkillsList")
print("2. EquipmentList")
print("3. StatusEffectsList")
print("4. PlayersList")
print("5. EnemiesList")
print("6. BossesList")

selectedSheet = int(input())
if (selectedSheet == 1):
	currentSheet = currentWorkbook["SkillsList"]
elif (selectedSheet == 2):
	currentSheet = currentWorkbook["EquipmentList"]
elif (selectedSheet == 3):
	currentSheet = currentWorkbook["StatusEffectsList"]
elif (selectedSheet == 4):
	currentSheet = currentWorkbook["PlayersList"]
elif (selectedSheet == 5):
	currentSheet = currentWorkbook["EnemiesList"]
elif (selectedSheet == 6):
	currentSheet = currentWorkbook["BossesList"]

if (currentSheet == currentWorkbook["SkillsList"]):
	print("\n1. Full")
	print("2. Name + Player")
	selectedMode = int(input())

	skillsAmount = int(input("\nAmount: "))

	skillsListString = "skillsList = {"
	for currentRow in range(2, skillsAmount + 3):
		currentID = int(currentSheet.cell(currentRow, 1).value)
		currentClass = currentSheet.cell(currentRow, 2).value
		if (currentClass in playerClassesValues):
			currentClass = str(playerClassesValues.index(currentClass))
		elif (currentClass in enemyNamesValues):
			currentClass = str(enemyNamesValues.index(currentClass))
		elif (currentClass in bossNamesValues):
			currentClass = str(bossNamesValues.index(currentClass))
		else:
			currentClass = "0"
		currentClassEval = eval(currentClass)
		currentName = currentSheet.cell(currentRow, 3).value
		currentMaxLevel = currentSheet.cell(currentRow, 4).value
		if (currentMaxLevel != "–"):
			currentMaxLevel = int(currentMaxLevel)
		else:
			currentMaxLevel = 0
		currentSkillType = currentSheet.cell(currentRow, 5).value
		for i in range(len(skillTypes)):
			if (currentSkillType == skillTypes[i]):
				currentSkillType = str(i)
				break
		currentStatType = currentSheet.cell(currentRow, 6).value
		for i in range(len(statTypes)):
			if (currentStatType == statTypes[i]):
				currentStatType = str(i)
				break
		currentElement = currentSheet.cell(currentRow, 7).value
		for i in range(len(elements)):
			if (currentElement == elements[i]):
				currentElement = str(i)
				break
		currentCost = currentSheet.cell(currentRow, 8).value
		if (currentCost != "–"):
			currentCost = int(currentCost)
		else:
			currentCost = 0
		currentTargetingTypeA = currentSheet.cell(currentRow, 9).value
		validTargetingTypeA = False
		for i in range(len(targetingTypes)):
			if (currentTargetingTypeA == targetingTypes[i] and targetingTypes[i] != "–"):
				currentTargetingTypeA = str(i)
				validTargetingTypeA = True
				break
		if (not validTargetingTypeA):
			currentTargetingTypeA = None
		currentTargetingTypeB = currentSheet.cell(currentRow, 10).value
		validTargetingTypeB = False
		for i in range(len(targetingTypes)):
			if (currentTargetingTypeB == targetingTypes[i] and targetingTypes[i] != "–"):
				currentTargetingTypeB = str(i)
				validTargetingTypeB = True
				break
		if (not validTargetingTypeB):
			currentTargetingTypeB = None
		currentBasePowerA = currentSheet.cell(currentRow, 11).value
		if (type(currentBasePowerA) is int or type(currentBasePowerA) is float):
			currentBasePowerA = "{:.2f}".format(currentBasePowerA)
		else:
			currentBasePowerA = None
		currentBasePowerB = currentSheet.cell(currentRow, 12).value
		if (type(currentBasePowerB) is int or type(currentBasePowerB) is float):
			currentBasePowerB = "{:.2f}".format(currentBasePowerB)
		else:
			currentBasePowerB = None
		currentAccuracyMod = currentSheet.cell(currentRow, 13).value
		if (currentAccuracyMod != "–"):
			currentAccuracyMod = "{:.2f}".format(currentAccuracyMod)
		else:
			currentAccuracyMod = "1.00"
		currentCritChance = currentSheet.cell(currentRow, 14).value
		if (currentCritChance != "–"):
			currentCritChance = "{:.2f}".format(currentCritChance)
		else:
			currentCritChance = "0.00"
		currentRandomnessMod = currentSheet.cell(currentRow, 15).value
		if (currentRandomnessMod != "–"):
			currentRandomnessMod = "{:.2f}".format(currentRandomnessMod)
		else:
			currentRandomnessMod = "0.10"
		currentStatusEffect = currentSheet.cell(currentRow, 16).value
		if (currentStatusEffect in statusEffectsValues):
			currentStatusEffect = str(statusEffectsValues.index(currentStatusEffect))
		else:
			currentStatusEffect = None
		currentStatusEffectDuration = currentSheet.cell(currentRow, 17).value
		if (currentStatusEffectDuration != "–"):
			currentStatusEffectDuration = int(currentStatusEffectDuration)
		else:
			currentStatusEffectDuration = 0

		if (selectedMode == 1):
			skillsListString += "\n\t" + str(currentID) + ": ["
			skillsListString += "\n\t\t" + str(currentClass) + ","
			skillsListString += '\n\t\t"' + str(currentName) + '",'
			skillsListString += "\n\t\t" + str(currentMaxLevel) + ","
			skillsListString += "\n\t\t" + str(currentSkillType) + ","
			skillsListString += "\n\t\t" + str(currentStatType) + ","
			skillsListString += "\n\t\t" + str(currentElement) + ","
			skillsListString += "\n\t\t" + str(currentCost) + ","
			skillsListString += "\n\t\t" + str(currentTargetingTypeA) + ","
			skillsListString += "\n\t\t" + str(currentTargetingTypeB) + ","
			skillsListString += "\n\t\t" + str(currentBasePowerA) + ","
			skillsListString += "\n\t\t" + str(currentBasePowerB) + ","
			skillsListString += "\n\t\t" + str(currentAccuracyMod) + ","
			skillsListString += "\n\t\t" + str(currentCritChance) + ","
			skillsListString += "\n\t\t" + str(currentRandomnessMod) + ","
			skillsListString += "\n\t\t" + str(currentStatusEffect) + ","
			skillsListString += "\n\t\t" + str(currentStatusEffectDuration)
			skillsListString += "\n\t],"
		elif (selectedMode == 2):
			skillsListString += "\n\t" + str(currentID) + ': ["' + str(currentName) + '", "' + str(currentClassEval) + '"],'
	skillsListString += "\n}"

	pyperclip.copy(skillsListString)

elif (currentSheet == currentWorkbook["EquipmentList"]):
	print("\n1. Full")
	print("2. Name")

	selectedMode = int(input())

	equipmentAmount = int(input("\nAmount: "))

	equipmentListString = "equipmentList = {\n"
	for currentRow in range(2, equipmentAmount + 3):
		currentID = int(currentSheet.cell(currentRow, 1).value)
		currentSlot = currentSheet.cell(currentRow, 2).value
		if (currentSlot == "Weapon"):
			currentSlot = 0
		elif (currentSlot == "Armor"):
			currentSlot = 1
		currentName = currentSheet.cell(currentRow, 3).value
		currentElement = currentSheet.cell(currentRow, 4).value
		currentElement = elementsValues.index(currentElement)
		currentBonusStats = []
		for i in range(8):
			currentStat = currentSheet.cell(currentRow, 5 + i).value
			if (currentStat == "–"):
				currentStat = "0"
			currentBonusStats.append(currentStat)

		if (selectedMode == 1):
			equipmentListString += "" + str(currentID) + ":Equipment("
			equipmentListString += "" + "equipmentSlots[" + str(currentSlot) + "],"
			equipmentListString += "" + '"' + str(currentName) + '",'
			equipmentListString += "" + "elements[" + str(currentElement) + "],"
			equipmentListString += "" + "Stats("
			for i in range(8):
				if (i != 0):
					equipmentListString += ","
				equipmentListString += str(currentBonusStats[i])
			equipmentListString += ")),"
		elif (selectedMode == 2):
			equipmentListString += "\n\t" + str(currentID) + ': "' + str(currentName) + '",'
	equipmentListString += "\n}"

	pyperclip.copy(equipmentListString)

elif (currentSheet == currentWorkbook["StatusEffectsList"]):
	print("\n1. Full")
	print("2. Name Only")
	selectedMode = int(input())

	statusEffectsAmount = int(input("\nAmount: "))

	statusEffectsListString = "statusEffectsList = {\n"
	for currentRow in range(2, statusEffectsAmount + 3):
		currentID = int(currentSheet.cell(currentRow, 1).value)
		currentName = currentSheet.cell(currentRow, 2).value
		currentHPBonus = currentSheet.cell(currentRow, 3).value
		if (currentHPBonus != "–"):
			currentHPBonus = int(currentHPBonus * 100)
		else:
			currentHPBonus = 0
		currentMPBonus = currentSheet.cell(currentRow, 4).value
		if (currentMPBonus != "–"):
			currentMPBonus = int(currentMPBonus * 100)
		else:
			currentMPBonus = 0
		currentMABonus = currentSheet.cell(currentRow, 5).value
		if (currentMABonus != "–"):
			currentMABonus = int(currentMABonus * 100)
		else:
			currentMABonus = 0
		currentRABonus = currentSheet.cell(currentRow, 6).value
		if (currentRABonus != "–"):
			currentRABonus = int(currentRABonus * 100)
		else:
			currentRABonus = 0
		currentMDBonus = currentSheet.cell(currentRow, 7).value
		if (currentMDBonus != "–"):
			currentMDBonus = int(currentMDBonus * 100)
		else:
			currentMDBonus = 0
		currentRDBonus = currentSheet.cell(currentRow, 8).value
		if (currentRDBonus != "–"):
			currentRDBonus = int(currentRDBonus * 100)
		else:
			currentRDBonus = 0
		currentACCBonus = currentSheet.cell(currentRow, 9).value
		if (currentACCBonus != "–"):
			currentACCBonus = int(currentACCBonus * 100)
		else:
			currentACCBonus = 0
		currentEVABonus = currentSheet.cell(currentRow, 10).value
		if (currentEVABonus != "–"):
			currentEVABonus = int(currentEVABonus * 100)
		else:
			currentEVABonus = 0

		if (selectedMode == 1):
			statusEffectsListString += "" + str(currentID) + ": StatusEffect("
			statusEffectsListString += '"' + str(currentName) + '",'
			statusEffectsListString += "Stats("
			statusEffectsListString += "" + str(currentHPBonus) + ","
			statusEffectsListString += "" + str(currentMPBonus) + ","
			statusEffectsListString += "" + str(currentMABonus) + ","
			statusEffectsListString += "" + str(currentRABonus) + ","
			statusEffectsListString += "" + str(currentMDBonus) + ","
			statusEffectsListString += "" + str(currentRDBonus) + ","
			statusEffectsListString += "" + str(currentACCBonus) + ","
			statusEffectsListString += "" + str(currentEVABonus) + ""
			statusEffectsListString += ")"
			statusEffectsListString += "),"
		elif (selectedMode == 2):
			statusEffectsListString += "\n\t" + str(currentID) + ': "' + str(currentName) + '",'
	statusEffectsListString += "\n}"

	pyperclip.copy(statusEffectsListString)

elif (currentSheet == currentWorkbook["PlayersList"]):
	# print("\n1. Name")
	# print("2. Data")
	# selectedMode = int(input())

	playersAmount = int(input("\nAmount: "))

	playersListString = "playersList = {"
	for currentRow in range(2, playersAmount + 3):
		currentID = int(currentSheet.cell(currentRow, 1).value)
		currentName = currentSheet.cell(currentRow, 2).value
		currentBaseStats = []
		for i in range(8):
			currentStat = currentSheet.cell(currentRow, 3 + i).value
			if (currentStat != "–"):
				currentStat = int(currentStat)
			else:
				currentStat = 1
			currentBaseStats.append(currentStat)
		currentSkills = []
		for i in range(9):
			currentSkill = currentSheet.cell(currentRow, 11 + i).value
			if (currentSkill in skillsNameValues):
				for i in skillsListKeys:
					if (skillsList[i][0] == currentSkill and skillsList[i][1] == currentName):
						currentSkills.append(str(skillsListKeys[i]))
						break
		currentWeapon = currentSheet.cell(currentRow, 20).value
		currentWeapon = equipmentValues.index(currentWeapon)
		currentArmor = currentSheet.cell(currentRow, 21).value
		currentArmor = equipmentValues.index(currentArmor)

		playersListString += "\n\t" + str(currentID) + ": Player("
		playersListString += '\n\t\t"' + str(currentName) + '",'
		playersListString += "\n\t\t["
		for i in range(8):
			if (i != 0):
				playersListString += ", "
			playersListString += str(currentBaseStats[i])
		playersListString += "],"
		playersListString += "\n\t\t["
		for i in range(len(currentSkills)):
			if (i != 0):
				playersListString += ", "
			playersListString += "skillsList[" + str(currentSkills[i]) + "]"
		playersListString += "],"
		playersListString += "\n\t\tequipmentList[" + str(currentWeapon) + "],"
		playersListString += "\n\t\tequipmentList[" + str(currentArmor) + "],"
		playersListString += '\n\t\t"050100050100"'
		playersListString += "\n\t),"
	playersListString += "\n}"

	pyperclip.copy(playersListString)

elif (currentSheet == currentWorkbook["EnemiesList"]):
	print("\n1. Name")
	print("2. Data")
	selectedMode = int(input())

	enemiesAmount = int(input("\nAmount: "))

	if (selectedMode == 1):
		enemyNamesString = "enemyNames = {"
		enemyNamesString += '\n\t0: "Nameless",'
	elif (selectedMode == 2):
		enemyDataString = "enemyData = {"
		enemyDataRegionString = '\t"region": {'
		enemyDataBaseStatsString = '\t"baseStats": {'
		enemyDataSkillsString = '\t"skills": {'
	for currentRow in range(2, enemiesAmount + 3):
		currentID = int(currentSheet.cell(currentRow, 1).value)
		currentName = currentSheet.cell(currentRow, 2).value
		currentRegion = currentSheet.cell(currentRow, 3).value
		if (currentRegion != "–"):
			currentRegion = int(currentRegion)
		else:
			currentRegion = 0
		currentBaseStats = []
		for i in range(8):
			currentStat = currentSheet.cell(currentRow, 4 + i).value
			if (currentStat != "–"):
				currentStat = int(currentStat)
			else:
				currentStat = 0
			if (i % 8 == 0 or i % 8 == 1):
				currentStat = "{:0>4}".format(currentStat)
			else:
				currentStat = "{:0>3}".format(currentStat)
			currentBaseStats.append(currentStat)
		currentBaseStats = "".join(currentBaseStats)
		currentSkills = []
		for i in range(9):
			currentSkill = currentSheet.cell(currentRow, 12 + i).value
			if (currentSkill == "Struggle"):
				for i in skillsListKeys:
					if (skillsList[i][0] == "Struggle"):
						currentSkills.append(str(skillsListKeys[i]))
						break
			elif (currentSkill in skillsNameValues):
				for i in skillsListKeys:
					if (skillsList[i][0] == currentSkill and skillsList[i][1] == currentName):
						currentSkills.append(str(skillsListKeys[i]))
						break
			else:
				currentSkills.append("0")
		for i in range(len(currentSkills)):
			currentSkills[i] = "{:0>3}".format(currentSkills[i])
		currentSkills = "".join(currentSkills)

		if (selectedMode == 1):
			enemyNamesString += "\n\t" + str(currentID) + ': "' + str(currentName) + '",'
		elif (selectedMode == 2):
			enemyDataRegionString += '\n\t\t' + str(currentID) + ': "' + str(currentRegion) + '",'
			enemyDataBaseStatsString += '\n\t\t' + str(currentID) + ': "' + str(currentBaseStats) + '",'
			enemyDataSkillsString += '\n\t\t' + str(currentID) + ': "' + str(currentSkills) + '",'
	if (selectedMode == 1):
		enemyNamesString += "\n}"
		pyperclip.copy(enemyNamesString)
	elif (selectedMode == 2):
		enemyDataRegionString += "\n\t},"
		enemyDataBaseStatsString += "\n\t},"
		enemyDataSkillsString += "\n\t}"
		enemyDataString += "\n"
		enemyDataString += enemyDataRegionString
		enemyDataString += "\n\n"
		enemyDataString += enemyDataBaseStatsString
		enemyDataString += "\n\n"
		enemyDataString += enemyDataSkillsString
		enemyDataString += "\n}"
		pyperclip.copy(enemyDataString)

elif (currentSheet == currentWorkbook["BossesList"]):
	print("\n1. Name")
	print("2. Data")
	selectedMode = int(input())

	bossesAmount = int(input("\nAmount: "))

	if (selectedMode == 1):
		bossNamesString = "bossNames = {"
		bossNamesString += '\n\t0: "Nameless",'
	elif (selectedMode == 2):
		bossDataString = "bossData = {"
		bossDataRegionString = '\t"region": {'
		bossDataBaseStatsString = '\t"baseStats": {'
		bossDataSkillsString = '\t"skills": {'
	for currentRow in range(2, bossesAmount + 3):
		currentID = int(currentSheet.cell(currentRow, 1).value)
		currentName = currentSheet.cell(currentRow, 2).value
		currentRegion = currentSheet.cell(currentRow, 3).value
		if (currentRegion != "–"):
			currentRegion = int(currentRegion)
		else:
			currentRegion = 0
		currentBaseStats = []
		for i in range(8):
			currentStat = currentSheet.cell(currentRow, 4 + i).value
			if (currentStat != "–"):
				currentStat = int(currentStat)
			else:
				currentStat = 0
			if (i % 8 == 0 or i % 8 == 1):
				currentStat = "{:0>4}".format(currentStat)
			else:
				currentStat = "{:0>3}".format(currentStat)
			currentBaseStats.append(currentStat)
		currentBaseStats = "".join(currentBaseStats)
		currentSkills = []
		for i in range(9):
			currentSkill = currentSheet.cell(currentRow, 12 + i).value
			if (currentSkill == "Struggle"):
				for i in skillsListKeys:
					if (skillsList[i][0] == "Struggle"):
						currentSkills.append(str(skillsListKeys[i]))
						break
			elif (currentSkill in skillsNameValues):
				for i in skillsListKeys:
					if (skillsList[i][0] == currentSkill and skillsList[i][1] == currentName):
						currentSkills.append(str(skillsListKeys[i]))
						break
			else:
				currentSkills.append("0")
		for i in range(len(currentSkills)):
			currentSkills[i] = "{:0>3}".format(currentSkills[i])
		currentSkills = "".join(currentSkills)

		if (selectedMode == 1):
			bossNamesString += "\n\t" + str(currentID) + ': "' + str(currentName) + '",'
		elif (selectedMode == 2):
			bossDataRegionString += '\n\t\t' + str(currentID) + ': "' + str(currentRegion) + '",'
			bossDataBaseStatsString += '\n\t\t' + str(currentID) + ': "' + str(currentBaseStats) + '",'
			bossDataSkillsString += '\n\t\t' + str(currentID) + ': "' + str(currentSkills) + '",'
	if (selectedMode == 1):
		bossNamesString += "\n}"
		pyperclip.copy(bossNamesString)
	elif (selectedMode == 2):
		bossDataRegionString += "\n\t},"
		bossDataBaseStatsString += "\n\t},"
		bossDataSkillsString += "\n\t}"
		bossDataString += "\n"
		bossDataString += bossDataRegionString
		bossDataString += "\n\n"
		bossDataString += bossDataBaseStatsString
		bossDataString += "\n\n"
		bossDataString += bossDataSkillsString
		bossDataString += "\n}"
		pyperclip.copy(bossDataString)

print("\nConversion complete.\n")