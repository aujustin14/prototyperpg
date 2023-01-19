import sys
import os
import random
import math
import openpyxl
import pyperclip

"""
todo:
	high priority
	- implement alt skills
	- implement skill leveling
	- implement equipment leveling
	- redo inventory and item system
	- redo enemy class
	- add more comments
	- improve enemy ai
	- balance damage, healing, and leveling
	- add more things to the town menu
	- implement saving/loading
	- create more items, item types, skills, and status effects
	- add more enemy/boss types
	- improve title screen
	low priority
"""


# Prints out a message and requires the user to press enter to continue. End determines where the cursor is placed (empty space is at the end of the string, \n is on a new line).
def proceduralPrint(string, end):
	print(string, end = end)
	input()


# Returns a bar based on a desired length, the alignment (0 for left align, 1 for right align), a character for the bar, a current value, and a max value.
def standardBar(length, alignment, characterString, currentValue, maxValue):
	if (len(characterString) > 1):
		characterString = characterString[0]

	if (alignment == 0):
		placeholderValue = "{:<" + str(length) + "}"
		borderStyle = "/"
	else:
		placeholderValue = "{:>" + str(length) + "}"
		borderStyle = "\\"

	barPercentage = currentValue / maxValue

	barDisplay = characterString * int(math.ceil((length * barPercentage)))

	return (borderStyle + placeholderValue.format(barDisplay) + borderStyle)


# Returns a row of the battle UI based on the provided string and the alignment (0 for left align, 1 for right align).
def battleUIRow(string, alignment):
	if (len(string) > 37):
		string = string[0:37]

	if (alignment == 0):
		currentRow = "| {:<37} |".format(string)
	else:
		currentRow = " {:>37} |".format(string)

	return currentRow


# Returns a row divider for the battle UI.
def battleUIRowDivider():
	return "+–––––––––––––––––––––––––––––––––––––––+–––––––––––––––––––––––––––––––––––––––+"


# Returns a row of the level up UI based on the provided value and the mode (0 for standard text, 1 for stat names, 2 for current stats, 3 for change in stats).
def levelUpUIRow(value, mode):
	currentRow = ""
	if (mode == 0):
		currentRow += "| " + str(value)
		currentRow += "\t" * (11 - math.ceil(len(currentRow) / 8))
		currentRow += "|"
	elif (mode == 1):
		currentRow += "|\tHP\tMP\tMA\tRA\tMD\tRD\tACC\tEVA\t\t|"
	elif (mode == 2):
		currentRow += "|"
		currentRow += "\t " + "{:q>4}".format(str(value.maxHP))
		currentRow += "\t " + "{:q>4}".format(str(value.maxMP))
		currentRow += "\t " + "{:q>4}".format(str(value.meleeAttack))
		currentRow += "\t " + "{:q>4}".format(str(value.rangedAttack))
		currentRow += "\t " + "{:q>4}".format(str(value.meleeDefense))
		currentRow += "\t " + "{:q>4}".format(str(value.rangedDefense))
		currentRow += "\t " + "{:q>4}".format(str(value.accuracy))
		currentRow += "\t " + "{:q>4}".format(str(value.evade))
		currentRow += "\t\t|"
		
		for i in range(len(currentRow)):
			if (currentRow[i] == "q"):
				currentRow = currentRow[0:i] + " " + currentRow[i+1:len(currentRow)]
	elif (mode == 3):
		currentRow += "|"
		currentRow += "\t+" + "{:q>4}".format(str(value.maxHP))
		currentRow += "\t+" + "{:q>4}".format(str(value.maxMP))
		currentRow += "\t+" + "{:q>4}".format(str(value.meleeAttack))
		currentRow += "\t+" + "{:q>4}".format(str(value.rangedAttack))
		currentRow += "\t+" + "{:q>4}".format(str(value.meleeDefense))
		currentRow += "\t+" + "{:q>4}".format(str(value.rangedDefense))
		currentRow += "\t+" + "{:q>4}".format(str(value.accuracy))
		currentRow += "\t+" + "{:q>4}".format(str(value.evade))
		currentRow += "\t\t|"

		for i in range(len(currentRow)):
			if (currentRow[i] == "q"):
				currentRow = currentRow[0:i] + " " + currentRow[i+1:len(currentRow)]

	return currentRow


# Rounds a value into an integer and then converts the result into another specified variable type.
def superRound(value, varType):
	return varType(round(value))


# Clears the screen.
def clearScreen():
	if (os.name == "nt"):
		os.system("cls")
	else:
		os.system("clear")


# Initialize classes #
class Stats:
	def __init__(self, maxHP=1, maxMP=1, meleeAttack=1, rangedAttack=1, meleeDefense=1, rangedDefense=1, accuracy=1, evade=1):
		self.maxHP = maxHP
		self.maxMP = maxMP
		self.meleeAttack = meleeAttack
		self.rangedAttack = rangedAttack
		self.meleeDefense = meleeDefense
		self.rangedDefense = rangedDefense
		self.accuracy = accuracy
		self.evade = evade
	
	# def __repr__(self):
	# 	return "HP: " + str(self.maxHP) + "\nMP: " + str(self.maxMP) + "\nM. Att.: " + str(self.meleeAttack) + "\nR. Att.: " + str(self.rangedAttack) + "\nM. Def.: " + str(self.meleeDefense) + "\nR. Def.: " + str(self.rangedDefense) + "\nAcc.: " + str(self.accuracy) + "\nEva.: " + str(self.evade)

	def __add__(self, other):
		if (type(other) is int):
			totalMaxHP = self.maxHP + other
			totalMaxMP = self.maxMP + other
			totalMeleeAttack = self.meleeAttack + other
			totalRangedAttack = self.rangedAttack + other
			totalMeleeDefense = self.meleeDefense + other
			totalRangedDefense = self.rangedDefense + other
			totalAccuracy = self.accuracy + other
			totalEvade = self.evade + other

		elif (type(other) is Stats):
			totalMaxHP = self.maxHP + other.maxHP
			totalMaxMP = self.maxMP + other.maxMP
			totalMeleeAttack = self.meleeAttack + other.meleeAttack
			totalRangedAttack = self.rangedAttack + other.rangedAttack
			totalMeleeDefense = self.meleeDefense + other.meleeDefense
			totalRangedDefense = self.rangedDefense + other.rangedDefense
			totalAccuracy = self.accuracy + other.accuracy
			totalEvade = self.evade + other.evade

		return Stats(
			totalMaxHP,
			totalMaxMP,
			totalMeleeAttack,
			totalRangedAttack,
			totalMeleeDefense,
			totalRangedDefense,
			totalAccuracy,
			totalEvade
		)

	def __sub__(self, other):
		if (type(other) is int):
			totalMaxHP = self.maxHP - other
			totalMaxMP = self.maxMP - other
			totalMeleeAttack = self.meleeAttack - other
			totalRangedAttack = self.rangedAttack - other
			totalMeleeDefense = self.meleeDefense - other
			totalRangedDefense = self.rangedDefense - other
			totalAccuracy = self.accuracy - other
			totalEvade = self.evade - other

		elif (type(other) is Stats):
			totalMaxHP = self.maxHP - other.maxHP
			totalMaxMP = self.maxMP - other.maxMP
			totalMeleeAttack = self.meleeAttack - other.meleeAttack
			totalRangedAttack = self.rangedAttack - other.rangedAttack
			totalMeleeDefense = self.meleeDefense - other.meleeDefense
			totalRangedDefense = self.rangedDefense - other.rangedDefense
			totalAccuracy = self.accuracy - other.accuracy
			totalEvade = self.evade - other.evade

		return Stats(
			totalMaxHP,
			totalMaxMP,
			totalMeleeAttack,
			totalRangedAttack,
			totalMeleeDefense,
			totalRangedDefense,
			totalAccuracy,
			totalEvade
		)


class Equipment:
	def __init__(self, slot, name, element, stats):
		self.slot = slot
		self.name = name
		self.element = element
		self.stats = stats
	
	# def __repr__(self):
	# 	return str(self.slot) + "\n" + str(self.name) + "\nLV " + str(self.level) + "\n" + str(self.element) + "\n" + str(self.stats)


class StatusEffect:
	def __init__(self, name, effect):
		self.name = name
		self.effect = effect

	def __repr__(self):
		return str(self.name)


class Skill:
	def __init__(self, forClass, name, maxPlayerLevel, skillType, statType, element, cost, targetingTypeA, targetingTypeB, basePowerA, basePowerB, accuracyMod, critChance, randomMod, statusEffect, statusEffectDuration):
		self.forClass = playerClasses[forClass]
		self.name = name
		self.currentLevel = 1
		self.maxPlayerLevel = maxPlayerLevel
		self.skillType = skillTypes[skillType]
		self.statType = statTypes[statType]
		self.element = elements[element]
		self.cost = cost
		if (targetingTypeA != None):
			self.targetingTypeA = targetingTypes[targetingTypeA]
		else:
			self.targetingTypeA = None
		if (targetingTypeB != None):
			self.targetingTypeB = targetingTypes[targetingTypeB]
		else:
			self.targetingTypeB = None
		self.basePowerA = basePowerA
		self.basePowerB = basePowerB
		self.accuracyMod = accuracyMod
		self.critChance = critChance
		self.randomMod = randomMod
		if (statusEffect != None):
			self.statusEffect = statusEffectsList[statusEffect]
		else:
			self.statusEffect = None
		self.statusEffectDuration = statusEffectDuration

	def __repr__(self):
		if (self.cost == 0):
			costString = "Free"
		else:
			costString = superRound(self.cost, str) + str(self.resource)

		targetingString = ""
		if (self.targetingTypeA != None):
			targetingString += self.targetingTypeA
		if (self.targetingTypeB != None):
			targetingString += "\t| " + self.targetingTypeB

		powerString = ""
		if (self.element in supportiveElements[1:3]):
			powerString += "–"
		else:
			if (self.basePowerA != None):
				powerString += superRound((self.basePowerA * 100), str) + "%"
			if (self.basePowerB != None):
				powerString += "\t\t| " + superRound((self.basePowerB * 100), str) + "%"
		
		if (self.accuracyMod >= 100):
			accuracyString = "Guaranteed"
		else:
			accuracyString = superRound((self.accuracyMod * 100), str) + "%"

		if (self.statusEffect != None and self.statusEffectDuration > 0):
			effectsString = str(self.statusEffect.name) + " for " + str(self.statusEffectDuration) + " turns"
		else:
			effectsString = "None"

		return (
			"+–––––––––––––––––––––––––––––––––––––––+"
			+ "\n| " + str(self.name)
			+ "\n+–––––––+–––––––––––––––––––––––––––––––+"
			+ "\n|Type\t| " + str(self.statType)
			+ "\n|Element| " + str(self.element)
			+ "\n|Target\t| " + targetingString
			+ "\n|Power\t| " + powerString
			+ "\n|Acc.\t| " + str(accuracyString)
			+ "\n|Effects| " + effectsString
			+ "\n|Cost\t| " + costString
			+ "\n+–––––––+–––––––––––––––––––––––––––––––+"
		)


class Item:
	def __init__(self, name, itemType, primaryPower, secondaryPower, value):
		self.name = name
		self.itemType = itemType
		self.primaryPower = primaryPower
		self.secondaryPower = secondaryPower
		self.value = value

	def __repr__(self):
		if (self.itemType == itemTypes[0]):
			powerString = str(self.primaryPower) + " HP or " + str(self.secondaryPower) + "% HP (whichever is higher)"
		else:
			powerString = str(self.primaryPower) + " MP or " + str(self.secondaryPower) + "% MP (whichever is higher)"

		return "\n+–––––––+–––––––––––––––––––––––––––––––+\n|Name\t| " + str(self.name) + "\n|Type\t| " + str(self.itemType) + "\n|Power\t| " + powerString + "\n+–––––––+–––––––––––––––––––––––––––––––+"


class Player:
	def __init__(self, currentName, currentBaseStats, currentSkills, currentWeapon, currentArmor, currentInventory):
		self.name = currentName
		self.currentHP = 1
		self.currentMP = 1

		self.baseStats = Stats(
			round(currentBaseStats[0]),
			round(currentBaseStats[1]),
			round(currentBaseStats[2]),
			round(currentBaseStats[3]),
			round(currentBaseStats[4]),
			round(currentBaseStats[5]),
			round(currentBaseStats[6]),
			round(currentBaseStats[7])
		)

		self.leveledUpStats = Stats(
			round(self.baseStats.maxHP * (1.05 ** (partyLevel - 1))),
			round(self.baseStats.maxMP * (1.01 ** (partyLevel - 1))),
			round(self.baseStats.meleeAttack * (1.05 ** (partyLevel - 1))),
			round(self.baseStats.rangedAttack * (1.05 ** (partyLevel - 1))),
			round(self.baseStats.meleeDefense * (1.05 ** (partyLevel - 1))),
			round(self.baseStats.rangedDefense * (1.05 ** (partyLevel - 1))),
			round(self.baseStats.accuracy * (1.05 ** (partyLevel - 1))),
			round(self.baseStats.evade * (1.05 ** (partyLevel - 1)))
		)

		self.skills = []
		for i in range(len(currentSkills)):
			skillData = skillsList[currentSkills[i]]
			self.skills.append(
				Skill(
					skillData[0],
					skillData[1],
					skillData[2],
					skillData[3],
					skillData[4],
					skillData[5],
					skillData[6],
					skillData[7],
					skillData[8],
					skillData[9],
					skillData[10],
					skillData[11],
					skillData[12],
					skillData[13],
					skillData[14],
					skillData[15]
				)
			)

		self.weapon = equipmentList[currentWeapon]
		self.armor = equipmentList[currentArmor]

		self.itemInventory = [
			int(currentInventory[0:2]),
			int(currentInventory[2:4]),
			int(currentInventory[4:6]),
			int(currentInventory[6:8]),
			int(currentInventory[8:10]),
			int(currentInventory[10:12])
		]

		self.statusEffects = []
		self.statusEffectDurations = []

		self.totalStats = Stats()
		self.evaluateTotalStats()

		self.currentHP = self.totalStats.maxHP
		self.currentMP = self.totalStats.maxMP

	def evaluateTotalStats(self):
		healthRatio = self.currentHP / self.totalStats.maxHP
		manaRatio = self.currentMP / self.totalStats.maxMP

		self.statusEffectStats = Stats(
			100,
			100,
			100,
			100,
			100,
			100,
			100,
			100
		)
		for i in range(len(self.statusEffects)):
			self.statusEffectStats += self.statusEffects[i].effect

		self.totalStats = Stats(
			round((self.leveledUpStats.maxHP + self.weapon.stats.maxHP + self.armor.stats.maxHP) * (self.statusEffectStats.maxHP / 100)),
			round((self.leveledUpStats.maxMP + self.weapon.stats.maxMP + self.armor.stats.maxMP) * (self.statusEffectStats.maxMP / 100)),
			round((self.leveledUpStats.meleeAttack + self.weapon.stats.meleeAttack + self.armor.stats.meleeAttack) * (self.statusEffectStats.meleeAttack / 100)),
			round((self.leveledUpStats.rangedAttack + self.weapon.stats.rangedAttack + self.armor.stats.rangedAttack) * (self.statusEffectStats.rangedAttack / 100)),
			round((self.leveledUpStats.meleeDefense + self.weapon.stats.meleeDefense + self.armor.stats.meleeDefense) * (self.statusEffectStats.meleeDefense / 100)),
			round((self.leveledUpStats.rangedDefense + self.weapon.stats.rangedDefense + self.armor.stats.rangedDefense) * (self.statusEffectStats.rangedDefense / 100)),
			round((self.leveledUpStats.accuracy + self.weapon.stats.accuracy + self.armor.stats.accuracy) * (self.statusEffectStats.accuracy / 100)),
			round((self.leveledUpStats.evade + self.weapon.stats.evade + self.armor.stats.evade) * (self.statusEffectStats.evade / 100))
		)

		self.currentHP = int(self.totalStats.maxHP * healthRatio)
		self.currentMP = int(self.totalStats.maxMP * manaRatio)

	def evaluateCurrentPoints(self):
		if (self.currentHP > self.totalStats.maxHP):
			self.currentHP = self.totalStats.maxHP
		elif (self.currentHP < 0):
			self.currentHP = 0

		if (self.currentMP > self.totalStats.maxMP):
			self.currentMP = self.totalStats.maxMP
		elif (self.currentMP < 0):
			self.currentMP = 0


class Enemy:
	def __init__(self, currentName, currentBaseStats, currentSkills):
		self.name = currentName
		self.currentClass = currentName
		# levelConstant = gameRegion * regionBattle
		self.currentLevel = random.choice([max(1, partyLevel - 1), max(1, partyLevel - 1), max(1, partyLevel - 1), partyLevel, partyLevel, partyLevel, partyLevel, min(maxPlayerLevel, partyLevel + 1)])
		self.currentHP = 1

		self.baseStats = Stats(
			round(currentBaseStats[0]),
			0,
			round(currentBaseStats[1]),
			round(currentBaseStats[2]),
			round(currentBaseStats[3]),
			round(currentBaseStats[4]),
			round(currentBaseStats[5]),
			round(currentBaseStats[6])
		)

		self.leveledUpStats = Stats(
			round(self.baseStats.maxHP * (1.05 ** (self.currentLevel - 1))),
			0,
			round(self.baseStats.meleeAttack * (1.05 ** (self.currentLevel - 1))),
			round(self.baseStats.rangedAttack * (1.05 ** (self.currentLevel - 1))),
			round(self.baseStats.meleeDefense * (1.05 ** (self.currentLevel - 1))),
			round(self.baseStats.rangedDefense * (1.05 ** (self.currentLevel - 1))),
			round(self.baseStats.accuracy * (1.05 ** (self.currentLevel - 1))),
			round(self.baseStats.evade * (1.05 ** (self.currentLevel - 1)))
		)

		# self.levelUpStats = Stats(
		# 	1.05,
		# 	1.05,
		# 	1.05,
		# 	1.05,
		# 	1.05,
		# 	1.05,
		# 	1.05
		# )

		self.skills = []
		for i in range(0, len(currentSkills), 3):
			skillData = skillsList[currentSkills[i]]
			self.skills.append(
				Skill(
					skillData[0],
					skillData[1],
					skillData[2],
					skillData[3],
					skillData[4],
					skillData[5],
					skillData[6],
					skillData[7],
					skillData[8],
					skillData[9],
					skillData[10],
					skillData[11],
					skillData[12],
					skillData[13],
					skillData[14],
					skillData[15]
				)
			)


		# self.currentWeapon = equipmentList[0]
		# self.currentArmor = equipmentList[1]

		self.statusEffects = []
		self.statusEffectDurations = []

		self.totalStats = Stats()
		self.evaluateTotalStats()
		# self.checkForLevelUp()

		self.currentHP = self.totalStats.maxHP

	def evaluateTotalStats(self):
		healthRatio = self.currentHP / self.totalStats.maxHP

		self.statusEffectStats = Stats(
			100,
			0,
			100,
			100,
			100,
			100,
			100,
			100
		)
		for i in range(len(self.statusEffects)):
			self.statusEffectStats += self.statusEffects[i].effect

		# self.totalStats = Stats(
		# 	superRound(self.baseStats.maxHP * (self.statusEffectStats.maxHP / 100), int),
		# 	superRound(self.baseStats.maxMP * (self.statusEffectStats.maxMP / 100), int),
		# 	superRound(self.baseStats.meleeAttack * (self.statusEffectStats.meleeAttack / 100), int),
		# 	superRound(self.baseStats.rangedAttack * (self.statusEffectStats.rangedAttack / 100), int),
		# 	superRound(self.baseStats.meleeDefense * (self.statusEffectStats.meleeDefense / 100), int),
		# 	superRound(self.baseStats.rangedDefense * (self.statusEffectStats.rangedDefense / 100), int),
		# 	superRound(self.baseStats.accuracy * (self.statusEffectStats.accuracy / 100), int),
		# 	superRound(self.baseStats.evade * (self.statusEffectStats.evade / 100), int)
		# )

		self.totalStats = Stats(
			round(self.leveledUpStats.maxHP * (self.statusEffectStats.maxHP / 100)),
			0,
			round(self.leveledUpStats.meleeAttack * (self.statusEffectStats.meleeAttack / 100)),
			round(self.leveledUpStats.rangedAttack * (self.statusEffectStats.rangedAttack / 100)),
			round(self.leveledUpStats.meleeDefense * (self.statusEffectStats.meleeDefense / 100)),
			round(self.leveledUpStats.rangedDefense * (self.statusEffectStats.rangedDefense / 100)),
			round(self.leveledUpStats.accuracy * (self.statusEffectStats.accuracy / 100)),
			round(self.leveledUpStats.evade * (self.statusEffectStats.evade / 100))
		)

		self.currentHP = superRound(self.totalStats.maxHP * healthRatio, int)

	def evaluateCurrentPoints(self):
		if (self.currentHP > self.totalStats.maxHP):
			self.currentHP = self.totalStats.maxHP
		elif (self.currentHP < 0):
			self.currentHP = 0

	# def checkForLevelUp(self):
	# 	self.baseStats = Stats(
	# 		math.ceil(self.baseStats.maxHP * (self.levelUpStats.maxHP ** (self.currentLevel - 1))),
	# 		math.ceil(self.baseStats.meleeAttack * (self.levelUpStats.meleeAttack ** (self.currentLevel - 1))),
	# 		math.ceil(self.baseStats.rangedAttack * (self.levelUpStats.rangedAttack ** (self.currentLevel - 1))),
	# 		math.ceil(self.baseStats.meleeDefense * (self.levelUpStats.meleeDefense ** (self.currentLevel - 1))),
	# 		math.ceil(self.baseStats.rangedDefense * (self.levelUpStats.rangedDefense ** (self.currentLevel - 1))),
	# 		math.ceil(self.baseStats.accuracy * (self.levelUpStats.accuracy ** (self.currentLevel - 1))),
	# 		math.ceil(self.baseStats.evade * (self.levelUpStats.evade ** (self.currentLevel - 1)))
	# 	)
	# 	self.evaluateTotalStats()


# Initialize game data #
maxPlayerLevel = 50

gameRegion = 1
regionBattle = 1
battleTurn = 0

partyMoney = 500
partyLevel = 1
partyCurrentXP = 0
partyNextXP = 200 * (1.15 ** (partyLevel - 1))
partyCurrentFocus = 1.00
partyMaxFocus = 1.00

currentPlayers = [None, None, None]
currentEnemies = [None, None, None]

selectedPlayer = -1
selectedEnemy = -1

playersAlive = [None, None, None]
enemiesAlive = [None, None, None]

playersHaveMoved = [None, None, None]
enemiesHaveMoved = [None, None, None]

playerInputInvalidCommand = False

playerClasses = {
	0: "None",
	1: "Class A",
	2: "Class B",
	3: "Class C",
}

enemyNames = {
	0: "Empty",
	1: "Green Slime",
	2: "Blue Slime",
	3: "Zombie",
	4: "Demon Eye",
	5: "Blood Slime",
	6: "Flesh Slime",
	7: "Man Eater",
	8: "Demon Eye Bat",
	9: "Stone Slime",
	10: "Flesh Slime",
	11: "Man Eater",
	12: "Demon Bat",
}

enemyData = {
	"region": {
		0: "0",
		1: "1",
		2: "1",
		3: "1",
		4: "1",
		5: "2",
		6: "2",
		7: "2",
		8: "2",
		9: "3",
		10: "3",
		11: "3",
		12: "3",
	},

	"baseStats": {
		0: "00000000000000000000000000",
		1: "02000050010009010009010010",
		2: "01500100009010009010010010",
		3: "02500025013008010005009007",
		4: "01500050012012006006010015",
		5: "02000100010010010010010010",
		6: "02500050012007011006010008",
		7: "02500025013009011006009008",
		8: "01500050013013007007010016",
		9: "02000100010010010010010010",
		10: "02500050012007011006010008",
		11: "02500025013009011006009008",
		12: "01500050013013007007010016",
	},

	"skills": {
		0: "001002001002001002001002001",
		1: "003004000000000000000000000",
		2: "005006000000000000000000000",
		3: "003004000000000000000000000",
		4: "005006000000000000000000000",
		5: "001002001002001002001002001",
		6: "001002001002001002001002001",
		7: "001002001002001002001002001",
		8: "001002001002001002001002001",
		9: "001002001002001002001002001",
		10: "001002001002001002001002001",
		11: "001002001002001002001002001",
		12: "001002001002001002001002001",
	}
}

bossNames = {
	0: "Empty",
	1: "Great Green Slime",
	2: "Great Blue Slime",
	3: "Armored Zombie",
	4: "Demon Eye Fleet",
	5: "Giant Blood Slime",
	6: "Giant Flesh Slime",
	7: "Stoic Man Eater",
	8: "Demon Bat Swarm",
}

bossData = {
	"region": {
		0: "0",
		1: "1",
		2: "1",
		3: "1",
		4: "1",
		5: "2",
		6: "2",
		7: "2",
		8: "2",
	},

	"baseStats": {
		0: "00000000000000000000000000",
		1: "03500100012011012011009009",
		2: "03000300011012011012009009",
		3: "04000025014009016011006006",
		4: "03000075010010010010008012",
		5: "03500100011011011011009009",
		6: "03500100013008012010009007",
		7: "04000025014010012007008007",
		8: "03000075014014008008008013",
	},

	"skills": {
		0: "000000000000000000000000000",
		1: "130131132000000000000000000",
		2: "133134135000000000000000000",
		3: "136137138000000000000000000",
		4: "139140141000000000000000000",
		5: "158159160161000000000000000",
		6: "162163164165000000000000000",
		7: "000000000000000",
		8: "000000000000000",
	}
}

regions = {
	0: "None",
	1: "Green Grasslands",
	2: "Crude Crimson",
	3: "Dark Depths",
}

equipmentSlots = {
	0: "Weapon",
	1: "Armor"
}

skillTypes = {
	0: "–",
	1: "Attack",
	2: "Heal",
	3: "Revive",
	4: "Buff",
	5: "Debuff",
}

statTypes = {
	0: "None",
	1: "Melee",
	2: "Ranged"
}

targetingTypes = {
	0: "None",
	1: "Self",
	2: "Single Enemy",
	3: "All Enemies",
	4: "Single Alive Ally",
	5: "All Alive Allies",
	6: "Single Dead Ally",
	7: "All Dead Allies",
}

elements = {
	0: "None",
	1: "Non-elemental",
	2: "Fire",
	3: "Water",
	4: "Ice",
	5: "Inherited",
	6: "Heal",
	7: "Revive",
	8: "Buff",
	9: "Debuff"
}

offensiveElements = [
	elements[1],
	elements[2],
	elements[3],
	elements[4],
	elements[5]
]

supportiveElements = [
	elements[6],
	elements[7],
	elements[8]
]

resources = {
	0: "None",
	1: " HP",
	2: " MP",
	3: " HP/MP",
	4: "% HP",
	5: "% MP",
	6: "% HP/MP"
}

hpResources = [
	resources[1],
	resources[4]
]

mpResources = [
	resources[2],
	resources[5]
]

flatResources = [
	resources[2],
	resources[3],
	resources[4]
]

percentageResources = [
	resources[5],
	resources[6],
	resources[6]
]

itemTypes = {
	0: "HP Recovery",
	1: "MP Recovery"
}

# Contains the list for status effects.
statusEffectsList = {
	0: StatusEffect("Healthy I",Stats(15,0,0,0,0,0,0,0)),1: StatusEffect("Healthy II",Stats(30,0,0,0,0,0,0,0)),2: StatusEffect("Healthy III",Stats(45,0,0,0,0,0,0,0)),3: StatusEffect("Energetic I",Stats(0,15,0,0,0,0,0,0)),4: StatusEffect("Energetic II",Stats(0,30,0,0,0,0,0,0)),5: StatusEffect("Energetic III",Stats(0,45,0,0,0,0,0,0)),6: StatusEffect("Strengthened I",Stats(0,0,15,0,0,0,0,0)),7: StatusEffect("Strengthened II",Stats(0,0,30,0,0,0,0,0)),8: StatusEffect("Strengthened III",Stats(0,0,45,0,0,0,0,0)),9: StatusEffect("Sharpened I",Stats(0,0,0,15,0,0,0,0)),10: StatusEffect("Sharpened II",Stats(0,0,0,30,0,0,0,0)),11: StatusEffect("Sharpened III",Stats(0,0,0,45,0,0,0,0)),12: StatusEffect("Shielded I",Stats(0,0,0,0,15,0,0,0)),13: StatusEffect("Shielded II",Stats(0,0,0,0,30,0,0,0)),14: StatusEffect("Shielded III",Stats(0,0,0,0,45,0,0,0)),15: StatusEffect("Barriered I",Stats(0,0,0,0,0,15,0,0)),16: StatusEffect("Barriered II",Stats(0,0,0,0,0,30,0,0)),17: StatusEffect("Barriered III",Stats(0,0,0,0,0,45,0,0)),18: StatusEffect("Accurate I",Stats(0,0,0,0,0,0,15,0)),19: StatusEffect("Accurate II",Stats(0,0,0,0,0,0,30,0)),20: StatusEffect("Accurate III",Stats(0,0,0,0,0,0,45,0)),21: StatusEffect("Evasive I",Stats(0,0,0,0,0,0,0,15)),22: StatusEffect("Evasive II",Stats(0,0,0,0,0,0,0,30)),23: StatusEffect("Evasive III",Stats(0,0,0,0,0,0,0,45)),24: StatusEffect("Resourceful I",Stats(15,15,0,0,0,0,0,0)),25: StatusEffect("Resourceful II",Stats(30,30,0,0,0,0,0,0)),26: StatusEffect("Resourceful III",Stats(45,45,0,0,0,0,0,0)),27: StatusEffect("Offensive I",Stats(0,0,15,15,0,0,15,0)),28: StatusEffect("Offensive II",Stats(0,0,30,30,0,0,30,0)),29: StatusEffect("Offensive III",Stats(0,0,45,45,0,0,45,0)),30: StatusEffect("Defensive I",Stats(0,0,0,0,15,15,0,15)),31: StatusEffect("Defensive II",Stats(0,0,0,0,30,30,0,30)),32: StatusEffect("Defensive III",Stats(0,0,0,0,45,45,0,45)),33: StatusEffect("Agile I",Stats(0,0,0,0,0,0,15,15)),34: StatusEffect("Agile II",Stats(0,0,0,0,0,0,30,30)),35: StatusEffect("Agile III",Stats(0,0,0,0,0,0,45,45)),36: StatusEffect("Enhanced I",Stats(15,0,15,0,15,0,0,0)),37: StatusEffect("Enhanced II",Stats(30,0,30,0,30,0,0,0)),38: StatusEffect("Enhanced III",Stats(45,0,45,0,45,0,0,0)),39: StatusEffect("Enchanted I",Stats(0,15,0,15,0,15,0,0)),40: StatusEffect("Enchanted II",Stats(0,30,0,30,0,30,0,0)),41: StatusEffect("Enchanted III",Stats(0,45,0,45,0,45,0,0)),42: StatusEffect("Empowered I",Stats(15,15,15,15,15,15,15,15)),43: StatusEffect("Empowered II",Stats(30,30,30,30,30,30,30,30)),44: StatusEffect("Empowered III",Stats(45,45,45,45,45,45,45,45)),45: StatusEffect("Ill I",Stats(-15,0,0,0,0,0,0,0)),46: StatusEffect("Ill II",Stats(-30,0,0,0,0,0,0,0)),47: StatusEffect("Ill III",Stats(-45,0,0,0,0,0,0,0)),48: StatusEffect("Lethargic I",Stats(0,-15,0,0,0,0,0,0)),49: StatusEffect("Lethargic II",Stats(0,-30,0,0,0,0,0,0)),50: StatusEffect("Lethargic III",Stats(0,-45,0,0,0,0,0,0)),51: StatusEffect("Weakened I",Stats(0,0,-15,0,0,0,0,0)),52: StatusEffect("Weakened II",Stats(0,0,-30,0,0,0,0,0)),53: StatusEffect("Weakened III",Stats(0,0,-45,0,0,0,0,0)),54: StatusEffect("Dulled I",Stats(0,0,0,-15,0,0,0,0)),55: StatusEffect("Dulled II",Stats(0,0,0,-30,0,0,0,0)),56: StatusEffect("Dulled III",Stats(0,0,0,-45,0,0,0,0)),57: StatusEffect("Broken I",Stats(0,0,0,0,-15,0,0,0)),58: StatusEffect("Broken II",Stats(0,0,0,0,-30,0,0,0)),59: StatusEffect("Broken III",Stats(0,0,0,0,-45,0,0,0)),60: StatusEffect("Shattered I",Stats(0,0,0,0,0,-15,0,0)),61: StatusEffect("Shattered II",Stats(0,0,0,0,0,-30,0,0)),62: StatusEffect("Shattered III",Stats(0,0,0,0,0,-45,0,0)),63: StatusEffect("Blinded I",Stats(0,0,0,0,0,0,-15,0)),64: StatusEffect("Blinded II",Stats(0,0,0,0,0,0,-30,0)),65: StatusEffect("Blinded III",Stats(0,0,0,0,0,0,-45,0)),66: StatusEffect("Sluggish I",Stats(0,0,0,0,0,0,0,-15)),67: StatusEffect("Sluggish II",Stats(0,0,0,0,0,0,0,-30)),68: StatusEffect("Sluggish III",Stats(0,0,0,0,0,0,0,-45)),69: StatusEffect("Drained I",Stats(-15,-15,0,0,0,0,0,0)),70: StatusEffect("Drained II",Stats(-30,-30,0,0,0,0,0,0)),71: StatusEffect("Drained III",Stats(-45,-45,0,0,0,0,0,0)),72: StatusEffect("Distracted I",Stats(0,0,-15,-15,0,0,-15,0)),73: StatusEffect("Distracted II",Stats(0,0,-30,-30,0,0,-30,0)),74: StatusEffect("Distracted III",Stats(0,0,-45,-45,0,0,-45,0)),75: StatusEffect("Offguard I",Stats(0,0,0,0,-15,-15,0,-15)),76: StatusEffect("Offguard II",Stats(0,0,0,0,-30,-30,0,-30)),77: StatusEffect("Offguard III",Stats(0,0,0,0,-45,-45,0,-45)),78: StatusEffect("Clumsy I",Stats(0,0,0,0,0,0,-15,-15)),79: StatusEffect("Clumsy II",Stats(0,0,0,0,0,0,-30,-30)),80: StatusEffect("Clumsy III",Stats(0,0,0,0,0,0,-45,-45)),81: StatusEffect("Poisoned I",Stats(-15,0,-15,0,-15,0,0,0)),82: StatusEffect("Poisoned II",Stats(-30,0,-30,0,-30,0,0,0)),83: StatusEffect("Poisoned III",Stats(-45,0,-45,0,-45,0,0,0)),84: StatusEffect("Cursed I",Stats(0,-15,0,-15,0,-15,0,0)),85: StatusEffect("Cursed II",Stats(0,-30,0,-30,0,-30,0,0)),86: StatusEffect("Cursed III",Stats(0,-45,0,-45,0,-45,0,0)),87: StatusEffect("Empowered I",Stats(-15,-15,-15,-15,-15,-15,-15,-15)),88: StatusEffect("Empowered II",Stats(-30,-30,-30,-30,-30,-30,-30,-30)),89: StatusEffect("Empowered III",Stats(-45,-45,-45,-45,-45,-45,-45,-45)),90: StatusEffect("Defending",Stats(0,0,0,0,50,50,0,0)),91: StatusEffect("Evading",Stats(0,0,0,0,0,0,0,50)),92: StatusEffect("Burn",Stats(-1,0,0,0,0,0,0,0)),93: StatusEffect("Blaze",Stats(-2,0,0,0,0,0,0,0)),94: StatusEffect("Scorch",Stats(-3,0,0,0,0,0,0,0)),95: StatusEffect("Wet",Stats(0,0,0,0,-10,-10,0,0)),96: StatusEffect("Soak",Stats(0,0,0,0,-20,-20,0,0)),97: StatusEffect("Drench",Stats(0,0,0,0,-30,-30,0,0)),98: StatusEffect("Frozen",Stats(0,0,0,0,0,0,-2,-2)),99: StatusEffect("Frostbite",Stats(0,0,0,0,0,0,-4,-4)),100: StatusEffect("Frostburn",Stats(0,0,0,0,0,0,-6,-6)),101: StatusEffect("Refreshing",Stats(0,0,0,0,0,0,0,0)),102: StatusEffect("Regrowing",Stats(0,0,0,0,0,0,0,0)),103: StatusEffect("Renewing",Stats(0,0,0,0,0,0,0,0)),104: StatusEffect("Napping",Stats(0,0,0,0,0,0,0,0)),105: StatusEffect("Resting",Stats(0,0,0,0,0,0,0,0)),106: StatusEffect("Relaxing",Stats(0,0,0,0,0,0,0,0)),
}

# Contains the list for fire-related status effects.
fireStatusEffects = [
	statusEffectsList[92],
	statusEffectsList[93],
	statusEffectsList[94]
]

# Contains the list for water-related status effects.
waterStatusEffects = [
	statusEffectsList[95],
	statusEffectsList[96],
	statusEffectsList[97]
]

# Contains the list for ice-related status effects.
iceStatusEffects = [
	statusEffectsList[98],
	statusEffectsList[99],
	statusEffectsList[100]
]

# Contains the list for HP-regen-related status effects.
hpRegenStatusEffects = [
	statusEffectsList[101],
	statusEffectsList[102],
	statusEffectsList[103]
]

# Contains the list for MP-regen-related status effects.
mpRegenStatusEffects = [
	statusEffectsList[104],
	statusEffectsList[105],
	statusEffectsList[106]
]

# Contains the list for equipment.
equipmentList = {
	0:Equipment(equipmentSlots[0],"Empty-handed",elements[0],Stats(0,0,0,0,0,0,0,0)),1:Equipment(equipmentSlots[1],"Naked",elements[0],Stats(0,0,0,0,0,0,0,0)),2:Equipment(equipmentSlots[0],"Sword",elements[1],Stats(0,0,5,0,0,0,0,0)),3:Equipment(equipmentSlots[1],"Chestplate",elements[1],Stats(0,0,0,0,5,0,0,0)),4:Equipment(equipmentSlots[0],"Dagger",elements[1],Stats(0,0,0,0,0,0,1,1)),5:Equipment(equipmentSlots[1],"Tunic",elements[1],Stats(0,0,0,0,0,0,1,1)),6:Equipment(equipmentSlots[0],"Staff",elements[1],Stats(0,0,0,5,0,0,0,0)),7:Equipment(equipmentSlots[1],"Robes",elements[1],Stats(0,0,0,0,0,5,0,0)),
}

# Contains the list for skills.
skillsList = {
	0: [
		0,
		"Empty",
		0,
		0,
		0,
		0,
		0,
		None,
		None,
		None,
		None,
		1.00,
		0.00,
		0.10,
		None,
		0
	],
	1: [
		0,
		"Defend",
		1,
		4,
		0,
		1,
		0,
		1,
		None,
		None,
		None,
		1.00,
		0.00,
		0.10,
		90,
		3
	],
	2: [
		0,
		"Evade",
		1,
		4,
		0,
		1,
		0,
		1,
		None,
		None,
		None,
		1.00,
		0.00,
		0.10,
		91,
		3
	],
	3: [
		1,
		"Basic Attack A",
		1,
		1,
		1,
		1,
		0,
		2,
		None,
		1.00,
		None,
		1.00,
		0.10,
		0.10,
		None,
		0
	],
	4: [
		1,
		"Basic Attack B",
		1,
		1,
		1,
		1,
		0,
		2,
		3,
		1.00,
		0.75,
		1.00,
		0.10,
		0.10,
		None,
		0
	],
	5: [
		2,
		"Basic Attack C",
		1,
		1,
		2,
		1,
		0,
		2,
		None,
		1.00,
		None,
		1.00,
		0.10,
		0.10,
		None,
		0
	],
	6: [
		2,
		"Basic Attack D",
		1,
		1,
		2,
		1,
		0,
		2,
		3,
		1.00,
		0.75,
		1.00,
		0.10,
		0.10,
		None,
		0
	],
	7: [
		3,
		"Self-heal",
		1,
		2,
		2,
		1,
		0,
		1,
		None,
		1.00,
		None,
		1.00,
		0.00,
		0.10,
		None,
		0
	],
	8: [
		3,
		"Heal A",
		1,
		2,
		2,
		1,
		0,
		4,
		None,
		1.00,
		None,
		1.00,
		0.00,
		0.10,
		None,
		0
	],
	9: [
		3,
		"Heal B",
		1,
		2,
		2,
		1,
		0,
		4,
		5,
		1.00,
		0.75,
		1.00,
		0.00,
		0.10,
		None,
		0
	],
	10: [
		3,
		"Revive A",
		1,
		3,
		2,
		1,
		0,
		6,
		None,
		1.00,
		None,
		1.00,
		0.00,
		0.10,
		None,
		0
	],
	11: [
		3,
		"Revive B",
		1,
		3,
		2,
		1,
		0,
		6,
		7,
		1.00,
		0.75,
		1.00,
		0.00,
		0.10,
		None,
		0
	],
}

# Contains the list for items.
itemsList = {
	0:Item("Lesser Health Potion",itemTypes[0],400,40,150),1:Item("Health Potion",itemTypes[0],2000,70,750),2:Item("Greater Health Potion",itemTypes[0],10000,100,3750),3:Item("Lesser Mana Potion",itemTypes[1],400,40,150),4:Item("Mana Potion",itemTypes[1],2000,70,750),5:Item("Greater Mana Potion",itemTypes[1],10000,100,3750)
}

# Contains the list for players.
playersList = {
	0: [
		"",
		[1, 1, 1, 1, 1, 1, 1, 1],
		[],
		0,
		1,
		"050100050100"
	],
	1: [
		"Player A",
		[550, 100, 55, 50, 55, 50, 20, 20],
		[3, 4],
		2,
		3,
		"050100050100"
	],
	2: [
		"Player B",
		[450, 100, 50, 50, 50, 50, 25, 25],
		[5, 6],
		4,
		5,
		"050100050100"
	],
	3: [
		"Player C",
		[500, 100, 50, 55, 50, 55, 20, 20],
		[7, 8, 9, 10, 11],
		6,
		7,
		"050100050100"
	],
}

# Contains the list for enemies.
enemiesList = {
	0: [
		"",
		[1, 1, 1, 1, 1, 1, 1],
		[]
	],
	1: [
		"Enemy A",
		[1500, 55, 50, 55, 50, 20, 20],
		[3, 4]
	],
	2: [
		"Enemy B",
		[750, 50, 50, 50, 50, 25, 25],
		[5, 6]
	],
	3: [
		"Enemy C",
		[1000, 50, 55, 50, 55, 20, 20],
		[7, 8, 9]
	],
}

# Contains the list for inn actions.
innActions = [
	"Eat",
	"Dine",
	"Rest",
	"Meditate",
	"Nap",
	"Slumber"
]

# Contains the list for inn power modifiers.
innPowers = [
	50,
	100,
	50,
	100,
	50,
	100
]

# Contains the list for inn recovery resources.
innResources = [
	resources[4],
	resources[4],
	resources[5],
	resources[5],
	resources[6],
	resources[6]
]

# Contains the list for costs of inn actions.
innCosts = [
	0,
	0,
	0,
	0,
	0,
	0
]

# Render the main menu of the game. This menu should give the player the option to start a new game, load a saved game, and quit the game.
def renderMainMenu():
	global playerInputInvalidCommand

	choice = -1

	while (choice != 0):
		clearScreen()
		if (playerInputInvalidCommand):
			print("+––          Invalid Command          ––+")
			playerInputInvalidCommand = False
		print("\nTitle")
		print("–––––")
		print("1. New Game")
		# print("2. Load Game")
		print("9. Debug")
		print("0. Exit")

		try:
			choice = int(input("> "))
			if (choice == 1):
				renderNewGameMenu()
			# elif (choice == 2):
				# renderLoadGameMenu()
			elif (choice == 9):
				renderDebugMenu()
			elif (choice == 0):
				sys.exit()
			else:
				playerInputInvalidCommand = True
		except ValueError:
			playerInputInvalidCommand = True


def renderDebugMenu():
	global skillsList

	currentLocation = "D:\MEGAsync Downloads\python rpg\python rpg.xlsx"
	currentWorkbook = openpyxl.load_workbook(filename = currentLocation, data_only = True)

	allPlayerNames = []
	for index in playersList:
		allPlayerNames.append(playersList[index][0])

	allEnemyNames = []
	for index in enemiesList:
		allEnemyNames.append(enemiesList[index][0])

	choice = -1

	while (choice != 0):
		clearScreen()
		print("\nDebug")
		print("–––––")
		print("1. Update Skills List")
		print("0. Back")

		try:
			choice = int(input("> "))
			if (choice == 1):
				# print(skillsList[0])
				skillsList = {}

				currentSheet = currentWorkbook["SkillsList"]

				currentRow = 2
				while ((currentSheet.cell(currentRow, 2)).value != None):
					currentSkillID = int(currentSheet.cell(currentRow, 1).value)

					currentSkillForEntity = currentSheet.cell(currentRow, 2).value
					if (currentSkillForEntity in allPlayerNames):
						currentSkillForEntity = allPlayerNames.index(currentSkillForEntity)
					elif (currentSkillForEntity in allEnemyNames):
						currentSkillForEntity = allEnemyNames.index(currentSkillForEntity)
					else:
						currentSkillForEntity = 0

					currentSkillName = currentSheet.cell(currentRow, 3).value

					currentSkillMaxLevel = currentSheet.cell(currentRow, 4).value
					if (currentSkillMaxLevel != "–"):
						currentSkillMaxLevel = int(currentSkillMaxLevel)
					else:
						currentSkillMaxLevel = 0

					currentSkillType = currentSheet.cell(currentRow, 5).value
					for i in range(len(skillTypes)):
						if (currentSkillType == skillTypes[i]):
							currentSkillType = i
							break
						if (i == (len(skillTypes) - 1)):
							currentSkillType = 0
							break

					currentSkillStatType = currentSheet.cell(currentRow, 6).value
					for i in range(len(statTypes)):
						if (currentSkillStatType == statTypes[i]):
							currentSkillStatType = i
							break
						if (i == (len(statTypes) - 1)):
							currentSkillStatType = 0
							break

					currentSkillElement = currentSheet.cell(currentRow, 7).value
					for i in range(len(elements)):
						if (currentSkillElement == elements[i]):
							currentSkillElement = i
							break
						if (i == (len(elements) - 1)):
							currentSkillElement = 0
							break

					currentSkillCost = currentSheet.cell(currentRow, 8).value
					if (currentSkillCost != "–"):
						currentSkillCost = int(currentSkillCost)
					else:
						currentSkillCost = 0

					currentTargetingTypeA = currentSheet.cell(currentRow, 9).value
					validTargetingTypeA = False
					for i in range(len(targetingTypes)):
						if (currentTargetingTypeA == targetingTypes[i] and targetingTypes[i] != "–"):
							currentTargetingTypeA = i
							validTargetingTypeA = True
							break
					if (not validTargetingTypeA):
						currentTargetingTypeA = None

					currentTargetingTypeB = currentSheet.cell(currentRow, 10).value
					validTargetingTypeB = False
					for i in range(len(targetingTypes)):
						if (currentTargetingTypeB == targetingTypes[i] and targetingTypes[i] != "–"):
							currentTargetingTypeB = i
							validTargetingTypeB = True
							break
					if (not validTargetingTypeB):
						currentTargetingTypeB = None

					currentBasePowerA = currentSheet.cell(currentRow, 11).value
					if (type(currentBasePowerA) is int or type(currentBasePowerA) is float):
						currentBasePowerA = float("{:.2f}".format(currentBasePowerA))
					else:
						currentBasePowerA = None

					currentBasePowerB = currentSheet.cell(currentRow, 12).value
					if (type(currentBasePowerB) is int or type(currentBasePowerB) is float):
						currentBasePowerB = float("{:.2f}".format(currentBasePowerB))
					else:
						currentBasePowerB = None

					currentAccuracyMod = currentSheet.cell(currentRow, 13).value
					if (currentAccuracyMod != "–"):
						currentAccuracyMod = float("{:.2f}".format(currentAccuracyMod))
					else:
						currentAccuracyMod = 1.00

					currentCritChance = currentSheet.cell(currentRow, 14).value
					if (currentCritChance != "–"):
						currentCritChance = float("{:.2f}".format(currentCritChance))
					else:
						currentCritChance = 0.00

					currentRandomnessMod = currentSheet.cell(currentRow, 15).value
					if (currentRandomnessMod != "–"):
						currentRandomnessMod = float("{:.2f}".format(currentRandomnessMod))
					else:
						currentRandomnessMod = 0.10

					currentStatusEffect = currentSheet.cell(currentRow, 16).value
					if (currentStatusEffect in statusEffectsList):
						currentStatusEffect = str(statusEffectsList.index(currentStatusEffect))
					else:
						currentStatusEffect = None

					currentStatusEffectDuration = currentSheet.cell(currentRow, 17).value
					if (currentStatusEffectDuration != "–"):
						currentStatusEffectDuration = int(currentStatusEffectDuration)
					else:
						currentStatusEffectDuration = 0
					
					skillsList[currentSkillID] = [
						currentSkillForEntity,
						currentSkillName,
						currentSkillMaxLevel,
						currentSkillType,
						currentSkillStatType,
						currentSkillElement,
						currentSkillCost,
						currentTargetingTypeA,
						currentTargetingTypeB,
						currentBasePowerA,
						currentBasePowerB,
						currentAccuracyMod,
						currentCritChance,
						currentRandomnessMod,
						currentStatusEffect,
						currentStatusEffectDuration
					]

					currentRow += 1

				print(skillsList)
				input()
				choice = 0
		except ValueError:
			continue


# Render the menu for when the player starts a new game. This menu should allow the player to check out and customize their character (such as their name and class). This menu should also allow the player to officially start a new game.
def renderNewGameMenu():
	global currentPlayers

	selectedCharacters = [1, 2, 3]

	choice = -1

	while (choice != 0):
		clearScreen()
		print("\nNew Game")
		print("–––––")
		if (not all(flag == 0 for flag in selectedCharacters)):
			print("1. Start")
		print("2. Change Character 1 (" + str(playersList[selectedCharacters[0]][0]) + ")")
		print("3. Change Character 2 (" + str(playersList[selectedCharacters[1]][0]) + ")")
		print("4. Change Character 3 (" + str(playersList[selectedCharacters[2]][0]) + ")")
		print("0. Back")

		try:
			choice = int(input("> "))
			if (choice == 1 and not all(flag == 0 for flag in selectedCharacters)):
				for i in range(len(currentPlayers)):
					if (selectedCharacters[i] != 0):
						characterData = playersList[selectedCharacters[i]]
						currentPlayers[i] = Player(characterData[0], characterData[1], characterData[2], characterData[3], characterData[4], characterData[5])
						playersAlive[i] = True
						playersHaveMoved[i] = False
				startGame()
				choice = 0
			elif (choice == 2):
				selectedCharacters[0] = requestCharacterChange(selectedCharacters[0], selectedCharacters)
			elif (choice == 3):
				selectedCharacters[1] = requestCharacterChange(selectedCharacters[1], selectedCharacters)
			elif (choice == 4):
				selectedCharacters[2] = requestCharacterChange(selectedCharacters[2], selectedCharacters)
		except ValueError:
			continue


# Render a menu that allows the player to change a character in the party.
def requestCharacterChange(currentCharacter, party):
	newCharacter = ""
	while (True):
		clearScreen()
		print("Change Character")
		print("–––––")
		print("Current Character: " + str(playersList[currentCharacter][0]))
		for key in playersList:
			if (key != 0 and key not in party):
				print(str(key) + ". " + str(playersList[key][0]))
		print("Q. Remove")
		print("0. Back")

		newCharacter = str(input("> "))
		if (newCharacter.isdecimal()):
			newCharacter = int(newCharacter)
			if (newCharacter != 0 and newCharacter in playersList):
				return int(newCharacter)
			elif (newCharacter == 0):
				return int(currentCharacter)
		else:
			if (newCharacter.upper() == "Q"):
				return 0


# Handles what happens inbetween battles and towns.
def startGame():
	global gameRegion
	global regionBattle
	global battleTurn

	gameRegion = 1
	regionBattle = 1

	while (playersAlive.count(True) > 0):
		if (regionBattle % 11 == 0):
			gameRegion += 1
			regionBattle = 1
			clearScreen()
			proceduralPrint("\nYou entered the " + str(regions[gameRegion % 3]) + ".", "")
		if (regionBattle % 5 == 0 or regionBattle % 5 == 1):
			clearScreen()
			proceduralPrint("\nYou came across a town.", "")
			renderTownActionMenu()
		startBattle()
		regionBattle += 1


# Create an enemy character for the player to fight. Also handles the battle system as well.
def startBattle():
	global currentPlayers
	global currentEnemies
	global selectedPlayer
	global selectedEnemy
	global playersHaveMoved
	global enemiesHaveMoved

	numOfEntities = random.randint(1, 3)
	for i in range(numOfEntities):
		entityData = enemiesList[random.randint(1, 3)]
		currentEnemies[i] = Enemy(entityData[0], entityData[1], entityData[2])
		enemiesAlive[i] = True
		enemiesHaveMoved[i] = False

	# if (regionBattle % 10 != 5 and regionBattle % 10 != 0):
	# 	numOfEnemies = random.randint(1, 3)
	# 	for i in range(numOfEnemies):
	# 		possibleEnemyKeys = []
	# 		for possibleEnemyKey in enemyData["region"]:
	# 			if (int(enemyData["region"][possibleEnemyKey]) == (gameRegion % 3)):
	# 				possibleEnemyKeys.append(possibleEnemyKey)
	# 		currentEnemyKey = random.choice(possibleEnemyKeys)
	# 		currentEnemyName = enemyNames[currentEnemyKey]
	# 		currentEnemyRegion = enemyData["region"][currentEnemyKey]
	# 		currentEnemyBaseStats = enemyData["baseStats"][currentEnemyKey]
	# 		currentEnemySkills = enemyData["skills"][currentEnemyKey]
	# 		currentEnemies[i] = Enemy(currentEnemyName, currentEnemyRegion, currentEnemyBaseStats, currentEnemySkills)
	# 		enemiesAlive[i] = True
	# 		enemiesHaveMoved[i] = False
	# else:
	# 	possibleBossKeys = []
	# 	for possibleBossKey in bossData["region"]:
	# 		if (int(bossData["region"][possibleBossKey]) == (gameRegion % 3)):
	# 			possibleBossKeys.append(possibleBossKey)
	# 	currentBossKey = random.choice(possibleBossKeys)
	# 	currentBossName = bossNames[currentBossKey]
	# 	currentBossRegion = bossData["region"][currentBossKey]
	# 	currentBossBaseStats = bossData["baseStats"][currentBossKey]
	# 	currentBossSkills = bossData["skills"][currentBossKey]
	# 	currentEnemies[0] = Enemy(currentBossName, currentBossRegion, currentBossBaseStats, currentBossSkills)
	# 	enemiesAlive[0] = True
	# 	enemiesHaveMoved[0] = False

	clearScreen()
	proceduralPrint("\nYou encountered a group of enemies!", "")
	progressBattleTurn()

	# While both the player and the enemy are alive, allow both characters to perform battle actions and progress the battle.
	while (playersAlive.count(True) > 0 and enemiesAlive.count(True) > 0):
		if (battleTurn % 2 == 1):
			while (playersHaveMoved.count(True) != (len(currentPlayers) - currentPlayers.count(None))):
				selectedPlayer = 0
				while (selectedPlayer < (len(currentPlayers) - currentPlayers.count(None))):
					if (playersHaveMoved[selectedPlayer] == True or playersHaveMoved[selectedPlayer] == None):
						selectedPlayer += 1
					else:
						break
				renderBattleActionMenu()
				evaluatePlayersEnemiesStatus()
		elif (battleTurn % 2 == 0):
			while (enemiesHaveMoved.count(True) != (len(currentEnemies) - currentEnemies.count(None))):
				selectedEnemy = 0
				while (selectedEnemy < (len(currentEnemies) - currentEnemies.count(None))):
					if (enemiesHaveMoved[selectedEnemy] == True or enemiesHaveMoved[selectedEnemy] == None):
						selectedEnemy += 1
					else:
						break
				initiateEnemyAttack()
				evaluatePlayersEnemiesStatus()
		progressBattleTurn()
	
	finishBattle()


# Determines what happens when a battle is finished.
def finishBattle():
	global partyCurrentXP
	global partyMoney

	clearScreen()

	renderBattleStatusMenu()

	if (playersAlive.count(True) == 0):
		proceduralPrint("\nYour party was killed.", "")
	elif (enemiesAlive.count(True) == 0):
		proceduralPrint("\nYou've killed your enemies.", "")

		gainedXP = 0
		for i in range(len(currentEnemies)):
			if (currentEnemies[i] != None):
				gainedXP += round(partyNextXP * (1 + 0.5 * (currentEnemies[i].currentLevel - partyLevel)))
		partyCurrentXP += gainedXP
		proceduralPrint("\nYour party gained " + str(gainedXP) + " XP.", "")

		gainedMoney = 0
		for i in range(len(currentEnemies)):
			if (currentEnemies[i] != None):
				gainedMoney = round(100 * (1.05 ** currentEnemies[i].currentLevel))
		partyMoney += gainedMoney
		proceduralPrint("\nYour party gained $" + str(gainedMoney) + ".", "")

		checkForPartyLevelUp()


# Checks if the party has leveled up. If true, increase each party member's stats.
def checkForPartyLevelUp():
	global partyLevel
	global partyCurrentXP
	global partyNextXP
	global innCosts

	partyLeveledUp = False

	initialPartyLevel = partyLevel
	initialStats = []
	healthRatio = []
	manaRatio = []
	for i in range(len(currentPlayers)):
		if (currentPlayers[i] != None):
			initialStats.append(currentPlayers[i].totalStats)
			healthRatio.append(currentPlayers[i].currentHP / currentPlayers[i].totalStats.maxHP)
			manaRatio.append(currentPlayers[i].currentMP / currentPlayers[i].totalStats.maxMP)

	while (partyCurrentXP >= partyNextXP and partyLevel < maxPlayerLevel):
		partyLeveledUp = True
		partyLevel += 1
		partyCurrentXP -= partyNextXP
		partyNextXP *= 1.15
		for i in range(len(currentPlayers)):
			if (currentPlayers[i] != None):
				currentPlayers[i].leveledUpStats = Stats(
					round(currentPlayers[i].baseStats.maxHP * (1.05 ** (partyLevel - 1))),
					round(currentPlayers[i].baseStats.maxMP * (1.01 ** (partyLevel - 1))),
					round(currentPlayers[i].baseStats.meleeAttack * (1.05 ** (partyLevel - 1))),
					round(currentPlayers[i].baseStats.rangedAttack * (1.05 ** (partyLevel - 1))),
					round(currentPlayers[i].baseStats.meleeDefense * (1.05 ** (partyLevel - 1))),
					round(currentPlayers[i].baseStats.rangedDefense * (1.05 ** (partyLevel - 1))),
					round(currentPlayers[i].baseStats.accuracy * (1.05 ** (partyLevel - 1))),
					round(currentPlayers[i].baseStats.evade * (1.05 ** (partyLevel - 1)))
				)
				currentPlayers[i].evaluateTotalStats()

	innCosts = [
		100 * partyLevel,
		200 * partyLevel,
		100 * partyLevel,
		200 * partyLevel,
		200 * partyLevel,
		400 * partyLevel
	]

	newPartyLevel = partyLevel
	newStats = []
	for i in range(len(currentPlayers)):
		if (currentPlayers[i] != None):
			newStats.append(currentPlayers[i].totalStats)

	if (partyLeveledUp):
		changeInStats = []
		for i in range(len(currentPlayers)):
			if (currentPlayers[i] != None):
				changeInStats.append(newStats[i] - initialStats[i])
				currentPlayers[i].currentHP = superRound(currentPlayers[i].totalStats.maxHP * healthRatio[i], int)
				currentPlayers[i].currentMP = superRound(currentPlayers[i].totalStats.maxMP * manaRatio[i], int)

		print("\n+–––––––––––––––––––––––––––––––––––––––––––––––––––––––––––––––––––––––––––––––+")
		print(levelUpUIRow("Level Up! LVL " + str(initialPartyLevel) + " -> LVL " + str(newPartyLevel), 0))
		print("+–––––––––––––––––––––––––––––––––––––––––––––––––––––––––––––––––––––––––––––––+")
		for i in range(3):
			if (currentPlayers[i] != None):
				print(levelUpUIRow(str(currentPlayers[i].name), 0))
				print(levelUpUIRow("", 1))
				print(levelUpUIRow(initialStats[i], 2))
				print(levelUpUIRow(changeInStats[i], 3))
			else:
				print(levelUpUIRow("", 0))
				print(levelUpUIRow("", 0))
				print(levelUpUIRow("", 0))
				print(levelUpUIRow("", 0))

			if (i < 2):
				print("|\t––––––––––––––––––––––––––––––––––––––––––––––––––––––––––––––––\t|")
		print("+–––––––––––––––––––––––––––––––––––––––––––––––––––––––––––––––––––––––––––––––+")
		proceduralPrint("", "")


# Render a menu that prints out the status of players and enemies when in battle (such as level, HP, and MP).
def renderBattleStatusMenu():
	global playerInputInvalidCommand

	clearScreen()

	if (partyCurrentFocus >= partyMaxFocus):
		playerFocusBar = "❇  " + superRound(partyCurrentFocus, str) + " />> MAX " + standardBar(30, 0, ">", 1, 1)[8:32]
	else:
		playerFocusBar = "❇  " + superRound(partyCurrentFocus, str) + " " + standardBar(30, 0, ">", (partyCurrentFocus % 1), 1.00)
	playerProgressText = str(regions[gameRegion % 3]) + " \\ Battle " + str(regionBattle) + " \\ Turn " + str(battleTurn)

	allPlayerTexts = []
	allPlayerBars = []
	allEnemyTexts = []
	allEnemyBars = []
	for i in range(3):
		if (currentPlayers[i] != None):
			currentPlayerName = currentPlayers[i].name
			currentPlayerCHP = currentPlayers[i].currentHP
			currentPlayerCMP = currentPlayers[i].currentMP
			currentPlayerMHP = currentPlayers[i].totalStats.maxHP
			currentPlayerMMP = currentPlayers[i].totalStats.maxMP

			if (len(currentPlayerName) > 14):
				currentPlayerName = currentPlayerName[0:11] + "..."
			# currentPlayerText = "/ " + "{:<14}".format(str(currentPlayerName)) + " / HP " + "{:^4}".format(str(currentPlayerCHP)) + " MP " + "{:^3}".format(str(currentPlayerCMP)) + " /"
			currentPlayerText = "/ " + "{:<14}".format(str(currentPlayerName)) + " / ❤  " + "{:^4}".format(str(currentPlayerCHP)) + " 🗲  " + "{:^3}".format(str(currentPlayerCMP)) + " /"
			if (i == selectedPlayer):
				currentPlayerText = "> " + currentPlayerText
			else:
				currentPlayerText = "  " + currentPlayerText
			currentPlayerBar = standardBar(22, 0, "=", currentPlayerCHP, currentPlayerMHP) + standardBar(10, 0, "—", currentPlayerCMP, currentPlayerMMP)
		else:
			currentPlayerText = ""
			currentPlayerBar = ""

		if (currentEnemies[i] != None and enemiesAlive[i]):
			currentEnemyName = "LV " + str(currentEnemies[i].currentLevel) + " " + currentEnemies[i].currentClass
			currentEnemyCHP = currentEnemies[i].currentHP
			currentEnemyMHP = currentEnemies[i].totalStats.maxHP

			if (len(currentEnemyName) > 20):
				currentEnemyName = currentEnemyName[0:17] + "..."
			# currentEnemyText = "\\ HP " + "{:^5}".format(str(currentEnemyCHP)) + " \\ " + "{:>20}".format(str(currentEnemyName)) + " \\"
			currentEnemyText = "\\ ❤  " + "{:^5}".format(str(currentEnemyCHP)) + " \\ " + "{:>20}".format(str(currentEnemyName)) + " \\"
			if (i == selectedEnemy):
				currentEnemyText += " <"
			else:
				currentEnemyText += "  "
			currentEnemyBar = standardBar(34, 1, "=", currentEnemyCHP, currentEnemyMHP)
		else:
			currentEnemyText = ""
			currentEnemyBar = ""

		allPlayerTexts.append(currentPlayerText)
		allPlayerBars.append(currentPlayerBar)
		allEnemyTexts.append(currentEnemyText)
		allEnemyBars.append(currentEnemyBar)

	print("\n", end="")
	print(battleUIRowDivider())

	print(battleUIRow(playerFocusBar, 0) + battleUIRow(playerProgressText, 1))

	print(battleUIRowDivider())

	print(battleUIRow(allPlayerTexts[0], 0) + battleUIRow(allEnemyTexts[0], 1))
	print(battleUIRow(allPlayerBars[0], 0) + battleUIRow(allEnemyBars[0], 1))
	# print("|\t–––––––––––––––––––––––––\t|\t–––––––––––––––––––––––––\t|")
	print(battleUIRow(allPlayerTexts[1], 0) + battleUIRow(allEnemyTexts[1], 1))
	print(battleUIRow(allPlayerBars[1], 0) + battleUIRow(allEnemyBars[1], 1))
	# print("|\t–––––––––––––––––––––––––\t|\t–––––––––––––––––––––––––\t|")
	print(battleUIRow(allPlayerTexts[2], 0) + battleUIRow(allEnemyTexts[2], 1))
	print(battleUIRow(allPlayerBars[2], 0) + battleUIRow(allEnemyBars[2], 1))

	print(battleUIRowDivider())
	print("|\t|\t|\t|\t|\t|\t|\t|\t|\t|\t|\t|")


# Render a menu that shows the player their general actions when in battle (such as use a skill or an item).
def renderBattleActionMenu():
	global playerInputInvalidCommand
	global battleTurn
	global selectedPlayer
	global currentPlayers
	global currentEnemies

	choice = -1

	while (choice != 0 and playersHaveMoved[selectedPlayer] != True):
		renderBattleStatusMenu()

		print("\n" + str(currentPlayers[selectedPlayer].name))
		print("–––––––")
		print("1. Attack")
		# print("1. Skills")
		print("2. Items")
		print("3. Scan")
		print("8. Defend")
		print("9. Evade")
		if (not (selectedPlayer == 0) and currentPlayers[0] != None and not playersHaveMoved[0]):
			print("Q. Switch to Player 1")
		if (not (selectedPlayer == 1) and currentPlayers[1] != None and not playersHaveMoved[1]):
			print("W. Switch to Player 2")
		if (not (selectedPlayer == 2) and currentPlayers[2] != None and not playersHaveMoved[2]):
			print("E. Switch to Player 3")
		print("0. Skip")

		try:
			choice = str(input("> "))
			if (choice == "1"):
				renderAttackMenu()
			elif (choice == "2"):
				renderItemMenu()
			elif (choice == "3"):
				renderScanMenu()
			elif (choice == "8"):
				defendSkillData = skillsList[1]
				castSkill([currentPlayers[selectedPlayer]], currentPlayers[selectedPlayer], Skill(defendSkillData[0],defendSkillData[1],defendSkillData[2],defendSkillData[3],defendSkillData[4],defendSkillData[5],defendSkillData[6],defendSkillData[7],defendSkillData[8],defendSkillData[9],defendSkillData[10],defendSkillData[11],defendSkillData[12],defendSkillData[13],defendSkillData[14],defendSkillData[15]), 0)
			elif (choice == "9"):
				evadeSkillData = skillsList[2]
				castSkill([currentPlayers[selectedPlayer]], currentPlayers[selectedPlayer], Skill(evadeSkillData[0],evadeSkillData[1],evadeSkillData[2],evadeSkillData[3],evadeSkillData[4],evadeSkillData[5],evadeSkillData[6],evadeSkillData[7],evadeSkillData[8],evadeSkillData[9],evadeSkillData[10],evadeSkillData[11],evadeSkillData[12],evadeSkillData[13],evadeSkillData[14],evadeSkillData[15]), 0)
			elif (choice.upper() == "Q" and not (selectedPlayer == 0) and currentPlayers[0] != None and not playersHaveMoved[0]):
				selectedPlayer = 0
			elif (choice.upper() == "W" and not (selectedPlayer == 1) and currentPlayers[1] != None and not playersHaveMoved[1]):
				selectedPlayer = 1
			elif (choice.upper() == "E" and not (selectedPlayer == 2) and currentPlayers[2] != None and not playersHaveMoved[2]):
				selectedPlayer = 2
			elif (choice == "0"):
				playersHaveMoved[selectedPlayer] = True
				proceduralPrint("\n" + str(currentPlayers[selectedPlayer].name) + " skipped their turn.", "")
			else:
				playerInputInvalidCommand = True
		except ValueError:
			playerInputInvalidCommand = True


# Render a menu that shows the player their available skills.
def renderAttackMenu():
	global playerInputInvalidCommand

	choice = -1

	while (choice != 0):
		renderBattleStatusMenu()
		print("\n" + str(currentPlayers[selectedPlayer].name) + " > Attack")
		print("––––––")
		for i in range(len(currentPlayers[selectedPlayer].skills)):
			count = i + 1
			currentSkill = currentPlayers[selectedPlayer].skills[i]

			currentSkillResource = currentPlayers[selectedPlayer].totalStats.maxMP
			resourceString = " MP"

			currentSkillTotalCost = currentSkill.cost

			if (currentSkill == skillsList[0]):
				continue
			elif (currentSkill.cost == 0):
				print(str(count) + ". " + currentSkill.name + " (Free)", end="\n")
			else:
				print(str(count) + ". " + currentSkill.name + " (" + str(currentSkillTotalCost) + str(resourceString) + ")")
		print("0. Back")

		try:
			choice = int(input("> "))
			if (choice >= 1 and choice <= len(currentPlayers[selectedPlayer].skills) and currentPlayers[selectedPlayer].skills[choice - 1] != skillsList[0]):
				choice -= 1
				chosenSkill = currentPlayers[selectedPlayer].skills[choice]

				currentPlayerResource = currentPlayers[selectedPlayer].currentMP
				chosenSkillResource = currentPlayers[selectedPlayer].totalStats.maxMP
				resourceString = "MP"

				chosenSkillCost = chosenSkill.cost

				if ((currentPlayerResource == currentPlayers[selectedPlayer].currentHP) and currentPlayerResource > chosenSkillCost):
					renderAttackDetails(chosenSkill)
					choice = 0
				elif ((currentPlayerResource == currentPlayers[selectedPlayer].currentMP) and currentPlayerResource >= chosenSkillCost):
					renderAttackDetails(chosenSkill)
					choice = 0
				else:
					playerInputInvalidCommand = True
					choice += 1
			elif (choice == 0):
				continue
			else:
				playerInputInvalidCommand = True
		except ValueError:
			playerInputInvalidCommand = True


# Render a menu that shows the player more details about the skill they just selected when in battle.
def renderAttackDetails(skill):
	global playerInputInvalidCommand

	choice = "-1"

	while (choice != "0"):
		renderBattleStatusMenu()
		print("\n" + str(currentPlayers[selectedPlayer].name) + " > Attack > Targeting")
		print(skill)

		targetAButtons = ["1", "2", "3"]
		if (skill.targetingTypeA == targetingTypes[1]):
			print(str(targetAButtons[0]) + ". Self")

		elif (skill.targetingTypeA == targetingTypes[2]):
			for i in range(len(currentEnemies) - enemiesAlive.count(None)):
				if (enemiesAlive[i]):
					print(targetAButtons[i] + ". " + str(currentEnemies[i].name))

		elif (skill.targetingTypeA == targetingTypes[3]):
			print(str(targetAButtons[0]) + ". All Enemies")

		elif (skill.targetingTypeA == targetingTypes[4]):
			for i in range(len(currentPlayers) - playersAlive.count(None)):
				if (playersAlive[i]):
					print(targetAButtons[i] + ". " + str(currentPlayers[i].name))

		elif (skill.targetingTypeA == targetingTypes[5]):
			print(str(targetAButtons[0]) + ". All Alive Allies")
		
		elif (skill.targetingTypeA == targetingTypes[6]):
			for i in range(len(currentPlayers) - playersAlive.count(None)):
				if (not playersAlive[i]):
					print(targetAButtons[i] + ". " + str(currentPlayers[i].name))

		elif (skill.targetingTypeA == targetingTypes[7]):
			if (playersAlive.count(False) >= 1):
				print(str(targetAButtons[0]) + ". All Dead Allies")

		elif (skill.targetingTypeA == None):
			pass

		else:
			print(str(targetAButtons[0]) + " ???")

		targetBButtons = ["Q", "W", "E"]
		if (skill.targetingTypeB == targetingTypes[1]):
			print(str(targetAButtons[0]) + " Self")

		elif (skill.targetingTypeB == targetingTypes[2]):
			for i in range(len(currentEnemies) - enemiesAlive.count(None)):
				if (enemiesAlive[i]):
					print(targetBButtons[i] + ". " + str(currentEnemies[i].name))

		elif (skill.targetingTypeB == targetingTypes[3]):
			print(str(targetBButtons[0]) + ". All Enemies")

		elif (skill.targetingTypeB == targetingTypes[4]):
			for i in range(len(currentPlayers) - playersAlive.count(None)):
				if (playersAlive[i]):
					print(targetBButtons[i] + ". " + str(currentPlayers[i].name))

		elif (skill.targetingTypeB == targetingTypes[5]):
			print(str(targetBButtons[0]) + ". All Alive Allies")
		
		elif (skill.targetingTypeB == targetingTypes[6]):
			for i in range(len(currentPlayers) - playersAlive.count(None)):
				if (not playersAlive[i]):
					print(targetBButtons[i] + ". " + str(currentPlayers[i].name))

		elif (skill.targetingTypeB == targetingTypes[7]):
			if (playersAlive.count(False) >= 1):
				print(str(targetBButtons[0]) + ". All Dead Allies")

		elif (skill.targetingTypeB == None):
			pass

		else:
			print(str(targetBButtons[0]) + ". ???")

		print("0. Cancel")

		try:
			choice = str(input("> ")).upper()[-1]
			if (len(choice) > 0 and choice in targetAButtons and skill.targetingTypeA != None):
				skillVersion = 0

				if (skill.targetingTypeA == targetingTypes[1]):
					if (choice == targetAButtons[0]):
						choice = targetAButtons.index(choice)
						chosenTarget = [currentPlayers[selectedPlayer]]
					else:
						choice = "-1"
						continue

				elif (skill.targetingTypeA == targetingTypes[2]):
					choice = targetAButtons.index(choice)
					if (enemiesAlive[choice]):
						chosenTarget = [currentEnemies[choice]]
					else:
						choice = "-1"
						continue

				elif (skill.targetingTypeA == targetingTypes[3]):
					if (choice == targetAButtons[0]):
						choice = targetAButtons.index(choice)
						chosenTarget = []
						for i in range(len(currentEnemies)):
							if (enemiesAlive[i] and currentEnemies[i] != None):
								chosenTarget.append(currentEnemies[i])
					else:
						choice = "-1"
						continue

				elif (skill.targetingTypeA == targetingTypes[4]):
					choice = targetAButtons.index(choice)
					if (playersAlive[choice]):
						chosenTarget = [currentPlayers[choice]]
					else:
						choice = "-1"
						continue

				elif (skill.targetingTypeA == targetingTypes[5]):
					if (choice == targetAButtons[0]):
						choice = targetAButtons.index(choice)
						chosenTarget = []
						for i in range(len(currentPlayers)):
							if (playersAlive[i] and currentPlayers[i] != None):
								chosenTarget.append(currentPlayers[i])
					else:
						choice = "-1"
						continue

				elif (skill.targetingTypeA == targetingTypes[6]):
					choice = targetAButtons.index(choice)
					if (not playersAlive[choice]):
						chosenTarget = [currentPlayers[choice]]
					else:
						choice = "-1"
						continue
				
				elif (skill.targetingTypeA == targetingTypes[7]):
					if (choice == targetAButtons[0]):
						choice = targetAButtons.index(choice)
						if (playersAlive.count(False) >= 1):
							chosenTarget = []
							for i in range(len(currentPlayers)):
								if (not playersAlive[i] and currentPlayers[i] != None):
									chosenTarget.append(currentPlayers[i])
						else:
							choice = "-1"
							continue
					else:
						choice = "-1"
						continue

				else:
					chosenTarget = []

			elif (len(choice) > 0 and choice in targetBButtons and skill.targetingTypeB != None):
				skillVersion = 1

				if (skill.targetingTypeB == targetingTypes[1]):
					if (choice == targetBButtons[0]):
						choice = targetBButtons.index(choice)
						chosenTarget = [currentPlayers[selectedPlayer]]
					else:
						choice = "-1"
						continue

				elif (skill.targetingTypeB == targetingTypes[2]):
					choice = targetBButtons.index(choice)
					if (enemiesAlive[choice]):
						chosenTarget = [currentEnemies[choice]]
					else:
						choice = "-1"
						continue

				elif (skill.targetingTypeB == targetingTypes[3]):
					if (choice == targetBButtons[0]):
						choice = targetBButtons.index(choice)
						chosenTarget = []
						for i in range(len(currentEnemies)):
							if (enemiesAlive[i] and currentEnemies[i] != None):
								chosenTarget.append(currentEnemies[i])
					else:
						choice = "-1"
						continue

				elif (skill.targetingTypeB == targetingTypes[4]):
					choice = targetBButtons.index(choice)
					if (playersAlive[choice]):
						chosenTarget = [currentPlayers[choice]]
					else:
						choice = "-1"
						continue

				elif (skill.targetingTypeB == targetingTypes[5]):
					if (choice == targetBButtons[0]):
						choice = targetBButtons.index(choice)
						chosenTarget = []
						for i in range(len(currentPlayers)):
							if (playersAlive[i] and currentPlayers[i] != None):
								chosenTarget.append(currentPlayers[i])
					else:
						choice = "-1"
						continue

				elif (skill.targetingTypeB == targetingTypes[6]):
					choice = targetBButtons.index(choice)
					if (not playersAlive[choice]):
						chosenTarget = [currentPlayers[choice]]
					else:
						choice = "-1"
						continue
				
				elif (skill.targetingTypeB == targetingTypes[7]):
					if (choice == targetBButtons[0]):
						choice = targetBButtons.index(choice)
						if (playersAlive.count(False) >= 1):
							chosenTarget = []
							for i in range(len(currentPlayers)):
								if (not playersAlive[i] and currentPlayers[i] != None):
									chosenTarget.append(currentPlayers[i])
						else:
							choice = "-1"
							continue
					else:
						choice = "-1"
						continue

				else:
					chosenTarget = []
			elif (len(choice) > 0 and choice == "0"):
				continue
			else:
				continue

			castSkill(chosenTarget, currentPlayers[selectedPlayer], skill, skillVersion)

			choice = "0"
		except:
			continue


# Render a menu that shows the player all the items.
def renderItemMenu():
	global playerInputInvalidCommand

	choice = -1

	while (choice != 0):
		renderBattleStatusMenu()
		print("\nItems")
		print("–––––")
		for i in range(len(itemsList)):
			currentItemAmount = currentPlayers[selectedPlayer].itemInventory[i]
			if (currentItemAmount > 0):
				print(str(i + 1) + ". ", end="")
				print(itemsList[i].name + " (" + str(currentItemAmount) + ")")
		print("0. Back")

		try:
			choice = int(input("> "))
			if (choice >= 1 and choice <= len(itemsList) and sum(currentPlayers[selectedPlayer].itemInventory) > 0):
				choice -= 1
				chosenItemAmount = currentPlayers[selectedPlayer].itemInventory[choice]

				if (chosenItemAmount > 0):
					renderItemDetails(choice)
					choice = 0
				else:
					playerInputInvalidCommand = True
					choice += 1
			elif (choice == 0):
				continue
			else:
				playerInputInvalidCommand = True
		except ValueError:
			playerInputInvalidCommand = True


# Render a menu that shows the player more details about the item they just selected when in battle.
def renderItemDetails(itemIndex):
	global playerInputInvalidCommand

	choice = -1
	chosenItem = itemsList[itemIndex]

	while (choice != 0):
		renderBattleStatusMenu()
		print("\nUse Item")
		print(chosenItem)
		print("1. Confirm")
		print("0. Back")

		try:
			choice = int(input("> "))
			if (choice == 1):
				useItem(currentPlayers[selectedPlayer], currentPlayers[selectedPlayer], itemIndex)
				choice = 0
			elif (choice == 0):
				continue
			else:
				playerInputInvalidCommand = True
		except ValueError:
			playerInputInvalidCommand = True


# Render a menu that shows the player who they can scan.
def renderScanMenu():
	global playerInputInvalidCommand

	choice = -1

	while (choice != 0):
		renderBattleStatusMenu()
		print("\nScan")
		print("–––––")
		print("1. " + str(currentPlayers[0].name))
		print("2. " + str(currentEnemies[0].name))
		print("0. Back")

		try:
			choice = int(input("> "))
			if (choice == 1):
				renderPlayerScanMenu()
				choice = 0
			elif (choice == 2):
				print(currentEnemies[0])
		except ValueError:
			playerInputInvalidCommand = True


# Render a menu that allows the player to scan themselves.
def renderPlayerScanMenu():
	global playerInputInvalidCommand

	choice = -1

	while (choice != 0):
		renderBattleStatusMenu()
		print("\nScan Player")
		print("–––––")
		print("1. Base Stats")
		print("2. Weapon")
		print("3. Armor")
		print("4. Status Effects")
		print("0. Back")

		try:
			choice = int(input("> "))
			if (choice == 1):
				proceduralPrint("\nBase Stats\n" + str(currentPlayers[0].baseStats), "\n")
				choice = 0
			elif (choice == 2):
				proceduralPrint("\n" + str(currentPlayers[0].weapon), "\n")
				choice = 0
			elif (choice == 3):
				proceduralPrint("\n" + str(currentPlayers[0].armor), "\n")
				choice = 0
			elif (choice == 4):
				statusEffectsString = ""
				if (len(currentPlayers[0].statusEffects) > 0):
					for i in range(len(currentPlayers[0].statusEffects)):
						statusEffectsString += "\n" + str(currentPlayers[0].statusEffects[i]) + " (" + str(currentPlayers[0].statusEffectDurations[i]) + ")"
						# statusEffectsString += str(currentPlayers[0].statusEffects[i])
						# statusEffectsString += " ("
						# statusEffectsString += str(currentPlayers[0].statusEffectDurations[i])
						# statusEffectsString += ")"
				else:
					statusEffectsString = "\nN/A"

				proceduralPrint(statusEffectsString, "\n")
				choice = 0
		except ValueError:
			playerInputInvalidCommand = True


# Have the enemy perform an action, such as casting a skill or using an item.
def initiateEnemyAttack():
	global currentPlayers
	global currentEnemies

	renderBattleStatusMenu()

	possibleSkills = []
	for possibleSkill in currentEnemies[selectedEnemy].skills:

		if (possibleSkill.name != "Empty" and possibleSkill.name != "Struggle"):
			possibleSkills.append(possibleSkill)

	# if (len(possibleSkills) == 0):
	# 	for currentKey in skillsList:
	# 		if (skillsList[currentKey].name == "Struggle"):
	# 			possibleSkills.append(str(skillsList[currentKey]))
	# 			break

	randomSkill = random.choice(possibleSkills)
	if (randomSkill.targetingTypeA == targetingTypes[1]):
		castSkill([currentEnemies[selectedEnemy]], currentEnemies[selectedEnemy], randomSkill, 0)

	elif (randomSkill.targetingTypeA == targetingTypes[2]):
		availableTargets = []
		for i in range(len(playersAlive)):
			if (playersAlive[i]):
				availableTargets.append(currentPlayers[i])
		randomTarget = random.choice(availableTargets)
		castSkill([randomTarget], currentEnemies[selectedEnemy], randomSkill, 0)

	elif (randomSkill.targetingTypeA == targetingTypes[3]):
		availableTargets = []
		for i in range(len(currentPlayers)):
			if (playersAlive[i] and currentPlayers[i] != None):
				availableTargets.append(currentPlayers[i])
		castSkill(availableTargets, currentEnemies[selectedEnemy], randomSkill, 0)

	elif (randomSkill.targetingTypeA == targetingTypes[4]):
		availableTargets = []
		for i in range(len(enemiesAlive)):
			if (enemiesAlive[i]):
				availableTargets.append(currentEnemies[i])
		randomTarget = random.choice(availableTargets)
		castSkill([randomTarget], currentEnemies[selectedEnemy], randomSkill, 0)

	elif (randomSkill.targetingTypeA == targetingTypes[5]):
		availableTargets = []
		for i in range(len(currentEnemies)):
			if (enemiesAlive[i] and currentEnemies[i] != None):
				availableTargets.append(currentEnemies[i])
		castSkill(availableTargets, currentEnemies[selectedEnemy], randomSkill, 0)

	else:
		castSkill([currentEnemies[selectedEnemy]], currentEnemies[selectedEnemy], randomSkill, 0)


# Check the status of all players and enemies.
def evaluatePlayersEnemiesStatus():
	global currentPlayers
	global currentEnemies
	global playersAlive
	global enemiesAlive

	for i in range(len(currentPlayers)):
		if (currentPlayers[i] != None):
			playersAlive[i] = currentPlayers[i].currentHP > 0
	for i in range(len(currentEnemies)):
		if (currentEnemies[i] != None):
			enemiesAlive[i] = currentEnemies[i].currentHP > 0


# Increment the turn in battle by 1.
def progressBattleTurn():
	global battleTurn
	global selectedPlayer
	global selectedEnemy
	global playersHaveMoved
	global enemiesHaveMoved

	battleTurn += 1

	if (battleTurn % 2 == 1):
		selectedPlayer = 0
		selectedEnemy = -1
	elif (battleTurn % 2 == 0):
		selectedPlayer = -1
		selectedEnemy = 0

	for i in range(len(playersHaveMoved)):
		if (playersAlive[i]):
			playersHaveMoved[i] = False
		elif (not playersAlive[i] and playersAlive[i] != None):
			playersHaveMoved[i] = True
	for i in range(len(enemiesHaveMoved)):
		if (enemiesAlive[i]):
			enemiesHaveMoved[i] = False
		elif (not enemiesAlive[i] and enemiesAlive[i] != None):
			enemiesHaveMoved[i] = True

	for i in range(len(currentPlayers)):
		if (currentPlayers[i] != None):
			for currentStatusEffect in currentPlayers[i].statusEffects:
				if (currentStatusEffect in fireStatusEffects):
					if (currentStatusEffect == fireStatusEffects[0]):
						currentPlayers[i].currentHP -= (2/100) * currentPlayers[i].totalStats.maxHP
					if (currentStatusEffect == fireStatusEffects[1]):
						currentPlayers[i].currentHP -= (4/100) * currentPlayers[i].totalStats.maxHP
					if (currentStatusEffect == fireStatusEffects[2]):
						currentPlayers[i].currentHP -= (6/100) * currentPlayers[i].totalStats.maxHP
		else:
			continue

	for i in range(len(currentEnemies)):
		if (currentEnemies[i] != None):
			for currentStatusEffect in currentEnemies[i].statusEffects:
				if (currentStatusEffect in fireStatusEffects):
					if (currentStatusEffect == fireStatusEffects[0]):
						currentEnemies[i].currentHP -= (2/100) * currentEnemies[i].totalStats.maxHP
					if (currentStatusEffect == fireStatusEffects[1]):
						currentEnemies[i].currentHP -= (4/100) * currentEnemies[i].totalStats.maxHP
					if (currentStatusEffect == fireStatusEffects[2]):
						currentEnemies[i].currentHP -= (6/100) * currentEnemies[i].totalStats.maxHP
		else:
			continue

	for i in range(len(currentPlayers)):
		if (currentPlayers[i] != None):
			expiredPlayerStatusEffects = []
			for j in range(len(currentPlayers[i].statusEffects)):
				currentPlayers[i].statusEffectDurations[j] -= 1
				if (currentPlayers[i].statusEffectDurations[j] == 0):
					expiredPlayerStatusEffects.append(j)
			expiredPlayerStatusEffects.reverse()
			for j in expiredPlayerStatusEffects:
				currentPlayers[i].statusEffects.pop(j)
				currentPlayers[i].statusEffectDurations.pop(j)
			currentPlayers[i].evaluateTotalStats()
		else:
			continue

	for i in range(len(currentEnemies)):
		if (currentEnemies[i] != None):
			expiredEnemyStatusEffects = []
			for j in range(len(currentEnemies[i].statusEffects)):
				currentEnemies[i].statusEffectDurations[j] -= 1
				if (currentEnemies[i].statusEffectDurations[j] == 0):
					expiredEnemyStatusEffects.append(j)
			expiredEnemyStatusEffects.reverse()
			for j in expiredEnemyStatusEffects:
				currentEnemies[i].statusEffects.pop(j)
				currentEnemies[i].statusEffectDurations.pop(j)
			currentEnemies[i].evaluateTotalStats()
		else:
			continue

	evaluatePlayersEnemiesStatus()


# Render a menu that prints out the status of the player when in town (such as level, HP, MP, and money).
def renderTownStatusMenu():
	global playerInputInvalidCommand

	clearScreen()
	if (playerInputInvalidCommand):
		print("+––          Invalid Command          ––+")
		playerInputInvalidCommand = False
	print("\n+–––––––––––––––––––––––––––––––––––––––+")
	print("| " + str(regions[gameRegion % 3]) + " | Town")
	print("+–––––––––––––––––––––––––––––––––––––––+")
	print("| " + str(currentPlayers[0].name) + " | LV " + str(partyLevel))
	print("| HP | " + str(currentPlayers[0].currentHP) + "/" + str(currentPlayers[0].totalStats.maxHP))
	print("| MP | " + str(currentPlayers[0].currentMP) + "/" + str(currentPlayers[0].totalStats.maxMP))
	print("| $$ | " + str(partyMoney))
	print("+–––––––––––––––––––––––––––––––––––––––+")


# Render a menu that shows the player their general actions when in town (go to the shop or enter the inn).
def renderTownActionMenu():
	global playerInputInvalidCommand

	choice = -1

	while (choice != 0):
		renderTownStatusMenu()
		print("\nAction")
		print("–––––")
		print("1. Shop")
		print("2. Inn")
		# print("9. Save Game")
		print("0. Leave")

		try:
			choice = int(input("> "))
			if (choice == 1):
				renderShopMenu()
			elif (choice == 2):
				renderInnMenu()
			elif (choice == 0):
				continue
			else:
				playerInputInvalidCommand = True
		except ValueError:
			playerInputInvalidCommand = True


# Render a menu that shows the player the shop in town.
def renderShopMenu():
	global playerInputInvalidCommand

	choice = -1

	while (choice != 0):
		renderTownStatusMenu()
		print("\nShop")
		print("–––––")
		for i in range(len(currentPlayers[0].itemInventory)):
			currentItemAmount = currentPlayers[0].itemInventory[i]
			print(str(i + 1) + ". ", end="")
			print(itemsList[i].name + " (" + str(currentItemAmount) + ") ($" + str(itemsList[i].value) + ")")
		print("0. Leave")

		try:
			choice = int(input("> "))
			if (choice >= 1 and choice <= len(itemsList)):
				choice -= 1
				chosenItemAmount = currentPlayers[0].itemInventory[choice]

				if (chosenItemAmount > 0):
					renderShopDetails(choice)
				else:
					playerInputInvalidCommand = True
			elif (choice == 0):
				continue
			else:
				playerInputInvalidCommand = True
		except ValueError:
			playerInputInvalidCommand = True


# Render a menu that shows the player more details about the item they just selected when in town.
def renderShopDetails(itemIndex):
	global playerInputInvalidCommand

	choice = -1
	chosenItem = itemsList[itemIndex]

	while (choice != 0):
		renderTownStatusMenu()
		print("\nBuy Item")
		print(chosenItem)
		print(">0. Buy current item at 100% value")
		print("<0. Sell current item at 20% value")
		print("0. Back")

		try:
			choice = int(input("> "))
			if (choice > 0):
				totalCost = chosenItem.value * choice
				if (partyMoney >= totalCost):
					partyMoney -= totalCost
					currentPlayers[0].itemInventory[itemIndex] += choice
					if (choice == 1):
						proceduralPrint("\n" + str(currentPlayers[0].name) + " spent $" + str(totalCost) + " to buy 1 " + str(chosenItem.name) + ".", "")
					else:
						proceduralPrint("\n" + str(currentPlayers[0].name) + " spent $" + str(totalCost) + " to buy " + str(choice) + " " + str(chosenItem.name) + "s.", "")
					choice = 0
				else:
					playerInputInvalidCommand = True
			elif (choice < 0):
				choice = abs(choice)
				if (currentPlayers[0].itemInventory[itemIndex] >= choice):
					totalProfit = (chosenItem.value * (20/100)) * choice
					totalProfit = superRound(totalProfit, int)
					partyMoney += totalProfit
					currentPlayers[0].itemInventory[itemIndex] -= choice
					if (choice == 1):
						proceduralPrint("\n" + str(currentPlayers[0].name) + " sold 1 " + str(chosenItem.name) + " for $" + str(totalProfit) + ".", "")
					else:
						proceduralPrint("\n" + str(currentPlayers[0].name) + " sold " + str(choice) + " " + str(chosenItem.name) + "s for $" + str(totalProfit) + ".", "")
					choice = 0
				else:
					playerInputInvalidCommand = True
			elif (choice == 0):
				continue
			else:
				playerInputInvalidCommand = True
		except ValueError:
			playerInputInvalidCommand = True


# Render a menu that shows the player the inn in town.
def renderInnMenu():
	global playerInputInvalidCommand

	choice = -1

	while (choice != 0):
		renderTownStatusMenu()
		print("\nInn")
		print("–––––")
		print("1. " + str(innActions[0]) + " (50% HP) ($" + str(innCosts[0]) + ")")
		print("2. " + str(innActions[1]) + " (100% HP) ($" + str(innCosts[1]) + ")")
		print("3. " + str(innActions[2]) + " (50% MP) ($" + str(innCosts[2]) + ")")
		print("4. " + str(innActions[3]) + " (100% MP) ($" + str(innCosts[3]) + ")")
		print("5. " + str(innActions[4]) + " (50% HP/MP) ($" + str(innCosts[4]) + ")")
		print("6. " + str(innActions[5]) + " (100% HP/MP) ($" + str(innCosts[5]) + ")")
		print("0. Leave")

		try:
			choice = int(input("> "))
			if (choice >= 1 and choice <= 6):
				choice -= 1
				chosenInnCosts = innCosts[choice]

				if (partyMoney >= chosenInnCosts):
					renderInnDetails(choice)
					choice = 0
				else:
					playerInputInvalidCommand = True
					choice += 1
			elif (choice == 0):
				continue
			else:
				playerInputInvalidCommand = True
		except ValueError:
			playerInputInvalidCommand = True


# Render a menu that shows the player more details about the inn action they just selected when in town.
def renderInnDetails(action):
	global playerInputInvalidCommand

	choice = -1

	while (choice != 0):
		renderTownStatusMenu()
		print("\nAction Details")
		print("+–––––––+–––––––––––––––––––––––––––––––+")
		print("|Action\t| " + str(innActions[action]))
		print("|Power\t| " + str(innPowers[action]) + str(innResources[action]))
		print("|Cost\t| " + str(innCosts[action]))
		print("+–––––––+–––––––––––––––––––––––––––––––+")
		print("\n1. Confirm")
		print("0. Cancel")

		try:
			choice = int(input("> "))
			if (choice == 1):
				chosenInnPower = innPowers[action]
				chosenInnResource = innResources[action]
				chosenInnCosts = innCosts[action]

				partyMoney -= chosenInnCosts

				if (chosenInnResource in flatResources):
					currentHPRecovery = superRound(chosenInnPower, int)
					currentMPRecovery = superRound(chosenInnPower, int)
				elif (chosenInnResource in percentageResources):
					currentHPRecovery = currentPlayers[0].totalStats.maxHP * (chosenInnPower / 100)
					currentMPRecovery = currentPlayers[0].totalStats.maxMP * (chosenInnPower / 100)
					currentHPRecovery = superRound(currentHPRecovery, int)
					currentMPRecovery = superRound(currentMPRecovery, int)
				else:
					currentHPRecovery = superRound(chosenInnPower, int)
					currentMPRecovery = superRound(chosenInnPower, int)

				if (chosenInnResource == resources[1] or chosenInnResource == resources[4]):
					currentPlayers[0].currentHP += currentHPRecovery
					currentPlayers[0].evaluateCurrentPoints()

					proceduralPrint("\n" + str(currentPlayers[0].name) + " has recovered " + str(currentHPRecovery) + " HP.", "")
					choice = 0
				elif (chosenInnResource == resources[2] or chosenInnResource == resources[5]):
					currentPlayers[0].currentMP += currentMPRecovery
					currentPlayers[0].evaluateCurrentPoints()

					proceduralPrint("\n" + str(currentPlayers[0].name) + " has recovered " + str(currentMPRecovery) + " MP.", "")
					choice = 0
				elif (chosenInnResource == resources[3] or chosenInnResource == resources[6]):
					currentPlayers[0].currentHP += currentHPRecovery
					currentPlayers[0].currentMP += currentMPRecovery
					currentPlayers[0].evaluateCurrentPoints()

					proceduralPrint("\n" + str(currentPlayers[0].name) + " has recovered " + str(currentHPRecovery) + " HP and " + str(currentMPRecovery) + " MP.", "")
					choice = 0
				else:
					continue
			elif (choice == 0):
				continue
			else:
				playerInputInvalidCommand = True
		except ValueError:
			playerInputInvalidCommand = True


# The user applies a specified status effect onto the target for a specified amount of turns.
def applyStatusEffect(target, user, statusEffect, statusEffectDuration):
	for i in range(len(target)):
		# If the target already has the status effect, reset the duration.
		if (statusEffect in target[i].statusEffects):
			statusEffectIndex = target[i].statusEffects.index(statusEffect)
			target[i].statusEffectDurations[statusEffectIndex] = statusEffectDuration + 1

		# Otherwise, the status effect is applied onto the target.
		else:
			target[i].statusEffects.append(statusEffect)
			target[i].statusEffectDurations.append(statusEffectDuration + 1)
		target[i].evaluateTotalStats()

		proceduralPrint(str(target[i].name) + " received " + str(statusEffect.name) + ".", "")


# The user casts a specified skill onto the target.
def castSkill(target, user, skill, version):
	global battleTurn
	global playersHaveMoved
	global enemiesHaveMoved

	clearScreen()
	renderBattleStatusMenu()

	# Get user text for strings
	userName = user.name
	targetName = []
	for i in range(len(target)):
		targetName.append(target[i].name)

	if (type(user) is Player):
		skillCost = skill.cost
		user.currentMP -= skillCost

	if (version == 0):
		skillPower = skill.basePowerA
	elif (version == 1):
		skillPower = skill.basePowerB
	else:
		skillPower = 1

	userAccuracyStat = user.totalStats.accuracy
	targetEvadeStat = []
	for i in range(len(target)):
		targetEvadeStat.append(target[i].totalStats.evade)

	proceduralPrint("\n" + userName + " casted " + skill.name + ".", "")

	if (skill.skillType == skillTypes[1]):
		damageMod = 1

		userAttackStat = None
		targetDefenseStat = []

		if (skill.statType == statTypes[1]):
			userAttackStat = user.totalStats.meleeAttack
			for i in range(len(target)):
				targetDefenseStat.append(target[i].totalStats.meleeDefense)
		elif (skill.statType == statTypes[2]):
			userAttackStat = user.totalStats.rangedAttack
			for i in range(len(target)):
				targetDefenseStat.append(target[i].totalStats.rangedDefense)
		else:
			userAttackStat = 1
			for i in range(len(target)):
				targetDefenseStat.append(1)

		if (skill.element == elements[5]):
			skill.element = user.weapon.element

		totalPotentialDamage = []
		totalHitRequirement = []
		totalCritRequirement = []
		for i in range(len(target)):
			currentPotentialDamage = math.ceil(4 * (userAttackStat ** 2) / (userAttackStat + targetDefenseStat[i]) * skillPower)
			if (currentPotentialDamage <= 0):
				currentPotentialDamage = 1
			totalPotentialDamage.append(currentPotentialDamage)

			hitRequirement = random.randint(0, 99)
			totalHitRequirement.append(hitRequirement)

			critRequirement = random.randint(0, 99)
			totalCritRequirement.append(critRequirement)

		currentAccuracy = []
		for i in range(len(target)):
			currentAccuracy.append(userAccuracyStat / targetEvadeStat[i] * skill.accuracyMod * 100)

		criticalAccuracy = []
		for i in range(len(target)):
			criticalAccuracy.append(skill.critChance * 100)

		totalDealtDamage = []
		for i in range(len(target)):
			currentDamage = round(totalPotentialDamage[i] * random.uniform((1.00 - (skill.randomMod / 2)), (1.00 + (skill.randomMod / 2))))

			if (currentAccuracy[i] > totalHitRequirement[i]):
				if (criticalAccuracy[i] > totalCritRequirement[i]):
					currentDamage = round(currentDamage * 1.5)
					criticalString = " critical"
				else:
					criticalString = ""

				target[i].currentHP -= currentDamage

				totalDealtDamage.append(currentDamage)
			else:
				totalDealtDamage.append(None)

		for i in range(len(target)):
			if (totalDealtDamage[i] != None):
				if (i == 0 and len(target) == 1):
					proceduralPrint(str(userName) + " dealt " + str(totalDealtDamage[i]) + str(criticalString) + " damage to " + str(targetName[i]) + ".", "")
				elif (i == 0 and len(target) > 1):
					print(str(userName) + " dealt " + str(totalDealtDamage[i]) + str(criticalString) + " damage to " + str(targetName[i]) + ",")
				elif (i > 0 and i < (len(target) - 1)):
					print((" " * len(userName)) + " dealt " + str(totalDealtDamage[i]) + str(criticalString) + " damage to " + str(targetName[i]) + ",")
				else:
					proceduralPrint((" " * len(userName)) + " dealt " + str(totalDealtDamage[i]) + str(criticalString) + " damage to " + str(targetName[i]) + ".", "")

				if (skill.statusEffect != None):
					applyStatusEffect(target[i], user, skill.statusEffect, skill.statusEffectDuration)
			else:
				if (i == 0 and len(target) == 1):
					proceduralPrint(str(userName) + " missed their attack on " + str(targetName[i]) + ".", "")
				elif (i == 0 and len(target) > 1):
					print(str(userName) + " missed their attack on " + str(targetName[i]) + ",")
				elif (i > 0 and i < (len(target) - 1)):
					print((" " * len(userName)) + " missed their attack on " + str(targetName[i]) + ",")
				else:
					proceduralPrint((" " * len(userName)) + " missed their attack on " + str(targetName[i]) + ".", "")

	elif (skill.skillType == skillTypes[2]):
		if (skill.statType == statTypes[1]):
			userHealStat = user.totalStats.meleeAttack
		elif (skill.statType == statTypes[2]):
			userHealStat = user.totalStats.rangedAttack
		else:
			userHealStat = 1

		totalPotentialHealing = []
		for i in range(len(target)):
			currentPotentialHealing = math.ceil((userHealStat ** 2) / (userHealStat) * skillPower)
			if (type(target[0]) is Player):
				currentPotentialHealing *= 4
			else:
				currentPotentialHealing *= 1.5
			if currentPotentialHealing < 0:
				currentPotentialHealing = 1
			totalPotentialHealing.append(currentPotentialHealing)

		totalDealtHealing = []
		for i in range(len(target)):
			currentHealing = round(totalPotentialHealing[i] * random.uniform((1.00 - (skill.randomMod / 2)), (1.00 + (skill.randomMod / 2))))

			target[i].currentHP += currentHealing

			totalDealtHealing.append(currentHealing)

		for i in range(len(target)):
			if (i == 0 and len(target) == 1):
				proceduralPrint(str(userName) + " dealt " + str(totalDealtHealing[i]) + " healing to " + str(targetName[i]) + ".", "")
			elif (i == 0 and len(target) > 1):
				print(str(userName) + " dealt " + str(totalDealtHealing[i]) + " healing to " + str(targetName[i]) + ",")
			elif (i > 0 and i < (len(target) - 1)):
				print((" " * len(userName)) + " dealt " + str(totalDealtHealing[i]) + " healing to " + str(targetName[i]) + ",")
			else:
				proceduralPrint((" " * len(userName)) + " dealt " + str(totalDealtHealing[i]) + " healing to " + str(targetName[i]) + ".", "")

		for i in range(len(target)):
			if (skill.statusEffect != None):
				applyStatusEffect(target[i], user, skill.statusEffect, skill.statusEffectDuration)

	elif (skill.skillType == skillTypes[3]):
		for i in range(len(target)):
			if (target[i].currentHP > 0):
				target.pop(target[i])
				targetName.pop(targetName[i])
				targetEvadeStat.pop(targetEvadeStat[i])

		if (skill.statType == statTypes[1]):
			userHealStat = user.totalStats.meleeAttack
		elif (skill.statType == statTypes[2]):
			userHealStat = user.totalStats.rangedAttack
		else:
			userHealStat = 1

		totalPotentialHealing = []
		for i in range(len(target)):
			currentPotentialHealing = round(((userHealStat ** 2) / (userHealStat)) * skillPower)
			if currentPotentialHealing < 0:
				currentPotentialHealing = 0
			totalPotentialHealing.append(currentPotentialHealing)

		totalDealtHealing = []
		for i in range(len(target)):
			currentHealing = round(totalPotentialHealing[i] * random.uniform((1.00 - (skill.randomMod / 2)), (1.00 + (skill.randomMod / 2))))

			target[i].currentHP += currentHealing
			if (user in currentPlayers):
				playersHaveMoved[currentPlayers.index(target[i])] = False
			elif (user in currentEnemies):
				enemiesHaveMoved[currentEnemies.index(target[i])] = False

			totalDealtHealing.append(currentHealing)

		for i in range(len(target)):
			if (i == 0 and len(target) == 1):
				proceduralPrint(str(userName) + " revived " + str(targetName[i]) + " with " +  str(totalDealtHealing[i]) + " healing.", "")
			elif (i == 0 and len(target) > 1):
				print(str(userName) + " revived " + str(targetName[i]) + " with " +  str(totalDealtHealing[i]) + " healing,")
			elif (i > 0 and i < (len(target) - 1)):
				print((" " * len(userName)) + " revived " + str(targetName[i]) + " with " +  str(totalDealtHealing[i]) + " healing,")
			else:
				proceduralPrint((" " * len(userName)) + " revived " + str(targetName[i]) + " with " +  str(totalDealtHealing[i]) + " healing.", "")

	elif (skill.skillType == skillTypes[4]):
		applyStatusEffect(target, user, skill.statusEffect, skill.statusEffectDuration)

	elif (skill.skillType == skillTypes[5]):
		hitChance = random.randint(1, 100)
		normalAccuracy = userAccuracyStat / targetEvadeStat * skill.accuracyMod * 100

		if (normalAccuracy >= hitChance):
			applyStatusEffect(target, user, skill.statusEffect, skill.statusEffectDuration)
		else:
			proceduralPrint(userName + " missed.", "")

	user.evaluateCurrentPoints()
	for i in range(len(target)):
		target[i].evaluateCurrentPoints()

	if (user in currentPlayers):
		playersHaveMoved[selectedPlayer] = True
	elif (user in currentEnemies):
		enemiesHaveMoved[selectedEnemy] = True


# The user uses an item onto the target (usually on itself).
def useItem(user, target, itemIndex):
	global battleTurn
	global playersHaveMoved
	global enemiesHaveMoved

	# Get user text for strings
	userName = user.name

	chosenItem = itemsList[itemIndex]

	# Reduce item amount by 1
	user.itemInventory[itemIndex] -= 1

	# String error checking
	if chosenItem.name[0].lower() in ["a", "e", "i", "o", "u"]:
		print("\n" + userName + " used an " + chosenItem.name + ".", end="")
	else:
		print("\n" + userName + " used a " + chosenItem.name + ".", end="")
	input()

	# If item is of health type
	if chosenItem.itemType == itemTypes[0]:
		currentHealing = max(chosenItem.primaryPower, int(round(((chosenItem.secondaryPower / 100) * user.totalStats.maxHP), 0)))
		user.currentHP += currentHealing
		print(userName + " recovered " + str(currentHealing) + " HP.", end="")

	# Else, if item is of mana type
	elif chosenItem.itemType == itemTypes[1]:
		currentHealing = max(chosenItem.primaryPower, int(round(((chosenItem.secondaryPower / 100) * user.totalStats.maxMP), 0)))
		user.currentMP += currentHealing
		print(userName + " recovered " + str(currentHealing) + " MP.", end="")

	# # Else, item is of damage type
	# else:
	# 	damageMod = 1

	# 	if skill.statType == statTypes[1]:
	# 		userAttackStat = user.totalStats.meleeAttack
	# 		targetDefenseStat = target.totalStats.meleeDefense
	# 	elif skill.statType == statTypes[2]:
	# 		userAttackStat = user.totalStats.rangedAttack
	# 		targetDefenseStat = target.totalStats.rangedDefense
	input()

	user.evaluateCurrentPoints()
	target.evaluateCurrentPoints()

	if (user in currentPlayers):
		playersHaveMoved[selectedPlayer] = True
	elif (user in currentEnemies):
		enemiesHaveMoved[selectedEnemy] = True


renderMainMenu()